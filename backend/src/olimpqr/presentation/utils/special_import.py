"""Parsers for the «Особая олимпиада» participant XLSX template.

The official template (`backend/templates/word/Участники.xlsx`) carries
positional data — column letters are stable but the header row can vary in
wording. This module recognizes the template by its header signature and
extracts participant fields, including the free-form passport / military
booklet block in column I.
"""

from __future__ import annotations

import datetime as dt
import hashlib
import io
import re
from typing import Any


_PASSPORT_BLOCK_RE = re.compile(r"\d{2}\s*\d{2}\s*[№N]?\s*\d{6}", re.UNICODE)
_DOB_RE = re.compile(r"дата\s+рождения\s+(\d{2}\.\d{2}\.\d{4})", re.IGNORECASE)
_DATE_RE = re.compile(r"(\d{2}\.\d{2}\.\d{4})")
_MILITARY_BOOKLET_RE = re.compile(
    r"(?:номер\s+военного\s+билета|военного\s+билета|военный\s+билет(?:а)?)[:\s]*"
    r"([А-ЯA-Z]{1,3}[\s№N]*\d+)",
    re.IGNORECASE | re.UNICODE,
)
_PERSONAL_NUMBER_RE = re.compile(
    r"личный\s+номер[:\s]*([A-Za-zА-Яа-я]{1,3}[\-\s]?\d+)",
    re.IGNORECASE | re.UNICODE,
)
_CITY_PAREN_RE = re.compile(r"\(\s*г\.?\s*([^)]+?)\s*\)", re.IGNORECASE | re.UNICODE)
_ISSUED_BY_PREFIX_RE = re.compile(r"^(?:кем\s+выдан|выдан(?:а|о)?)\s+", re.IGNORECASE | re.UNICODE)
_MIL_BOOKLET_TAIL_RE = re.compile(
    r"(?:номер\s+военного|военного\s+билета|военный\s+билет|личный\s+номер)",
    re.IGNORECASE | re.UNICODE,
)


def _coerce(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_ru_date(text: str) -> dt.date | None:
    text = (text or "").strip()
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def split_institution_and_city(raw: str) -> tuple[str, str | None]:
    """Strip a `(г. <city>)` suffix from an institution name.

    Returns (institution_name_without_city, city_or_None).
    """
    raw = _coerce(raw)
    if not raw:
        return "", None
    match = _CITY_PAREN_RE.search(raw)
    city: str | None = None
    if match:
        city = match.group(1).strip()
        raw = (raw[: match.start()] + raw[match.end():]).strip()
    raw = re.sub(r"\s{2,}", " ", raw).strip(" ,;-")
    return raw, city or None


def parse_passport_military_block(text: str) -> dict[str, Any]:
    """Extract structured fields from the free-form column I cell."""
    text = _coerce(text)
    out: dict[str, Any] = {
        "dob": None,
        "passport_series_number": None,
        "passport_issued_by": None,
        "passport_issued_date": None,
        "military_booklet_number": None,
        "military_personal_number": None,
    }
    if not text:
        return out

    dob_match = _DOB_RE.search(text)
    if dob_match:
        out["dob"] = _parse_ru_date(dob_match.group(1))
        text_after_dob = text[dob_match.end():]
    else:
        text_after_dob = text

    passport_match = _PASSPORT_BLOCK_RE.search(text_after_dob)
    if passport_match:
        out["passport_series_number"] = re.sub(r"\s+", " ", passport_match.group(0)).strip()
        # Everything between the passport series and the next date is the
        # issuing authority. The issue date is that next date itself.
        tail = text_after_dob[passport_match.end():]
        # Stop at the start of the military booklet block, if any.
        tail_stop = _MIL_BOOKLET_TAIL_RE.search(tail)
        if tail_stop:
            tail = tail[: tail_stop.start()]
        date_match = _DATE_RE.search(tail)
        if date_match:
            out["passport_issued_date"] = _parse_ru_date(date_match.group(1))
            issued_by = tail[: date_match.start()]
        else:
            issued_by = tail
        issued_by = issued_by.strip(" ,;:-.")
        issued_by = _ISSUED_BY_PREFIX_RE.sub("", issued_by).strip(" ,;:-.")
        issued_by = re.sub(r"\s{2,}", " ", issued_by)
        if issued_by:
            out["passport_issued_by"] = issued_by

    booklet_match = _MILITARY_BOOKLET_RE.search(text)
    if booklet_match:
        out["military_booklet_number"] = re.sub(r"\s+", " ", booklet_match.group(1)).strip()

    personal_match = _PERSONAL_NUMBER_RE.search(text)
    if personal_match:
        out["military_personal_number"] = personal_match.group(1).strip()

    return out


def derive_imported_email(full_name: str, dob: dt.date | None) -> str:
    """Stable synthetic email so that re-importing updates instead of duplicating."""
    payload = f"{(full_name or '').strip().lower()}|{dob.isoformat() if dob else ''}"
    digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]
    return f"imported.{digest}@participants.local"


def looks_like_uchastniki_sheet(rows: list[tuple[Any, ...]]) -> bool:
    """Detect the «Участники» template by header signature."""
    if not rows:
        return False
    header = [_coerce(c).lower() for c in rows[0]]
    if len(header) < 9:
        return False
    inst_cell = header[1] if len(header) > 1 else ""
    fio_cell = header[6] if len(header) > 6 else ""
    passport_cell = header[8] if len(header) > 8 else ""
    inst_ok = any(token in inst_cell for token in ("воен", "вуз", "учеб", "наименование"))
    fio_ok = "фио" in fio_cell
    passport_ok = "паспорт" in passport_cell
    return inst_ok and fio_ok and passport_ok


def parse_uchastniki_xlsx(file_bytes: bytes) -> list[dict[str, Any]] | None:
    """Try to parse the file as a «Участники» template.

    Returns a list of normalized records (one per data row) when the template
    is recognized, or `None` if the file does not match — caller should fall
    back to the legacy header-based parser.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("Для импорта XLSX требуется зависимость openpyxl") from exc

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not looks_like_uchastniki_sheet(rows):
            continue
        records: list[dict[str, Any]] = []
        for row in rows[1:]:
            if not any(cell is not None and _coerce(cell) for cell in row):
                continue
            # Normalize tuple length so positional access never raises
            row_padded = list(row) + [None] * (11 - len(row))
            full_name = _coerce(row_padded[6])
            if not full_name:
                continue
            institution_raw = _coerce(row_padded[1])
            institution_clean, city = split_institution_and_city(institution_raw)
            position = _coerce(row_padded[4]) or None
            military_rank = _coerce(row_padded[5]) or None
            captain_cell = _coerce(row_padded[7])
            is_captain = bool(captain_cell)
            block = parse_passport_military_block(_coerce(row_padded[8]))

            record: dict[str, Any] = {
                "full_name": full_name,
                "institution": institution_clean or institution_raw,
                "institution_location": city,
                "is_captain": is_captain,
                "position": position,
                "military_rank": military_rank,
                "dob": block["dob"],
                "passport_series_number": block["passport_series_number"],
                "passport_issued_by": block["passport_issued_by"],
                "passport_issued_date": block["passport_issued_date"],
                "military_booklet_number": block["military_booklet_number"],
                "military_personal_number": block["military_personal_number"],
            }
            records.append(record)
        return records
    return None
