"""Admin API endpoints."""

import csv
import io
import json
import re
import secrets
import zipfile
from urllib.parse import quote
from html import escape
from datetime import datetime
from pathlib import Path
from typing import Annotated, Any, Optional
from uuid import UUID, uuid4
from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, Response, UploadFile, status
from fastapi.responses import HTMLResponse, StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import delete, insert, select, update
from sqlalchemy.orm import selectinload

from ....infrastructure.database import get_db
from ....infrastructure.database.models import BadgeTemplateModel, RoomModel, SeatAssignmentModel, TourTimeModel
from ....infrastructure.repositories import (
    UserRepositoryImpl,
    AuditLogRepositoryImpl,
    CompetitionRepositoryImpl,
    ScanRepositoryImpl,
    RegistrationRepositoryImpl,
    ParticipantRepositoryImpl,
    EntryTokenRepositoryImpl,
    InstitutionRepositoryImpl,
    AttemptRepositoryImpl,
    AnswerSheetRepositoryImpl,
    RoomRepositoryImpl,
    SeatAssignmentRepositoryImpl,
    UserCompetitionAccessRepositoryImpl,
)
from ....infrastructure.security import hash_password
from ....infrastructure.storage import MinIOStorage
from ....infrastructure.pdf import SheetGenerator
from ....infrastructure.docx import WordTemplateGenerator
from ....domain.entities import SeatAssignment, User
from ....domain.value_objects import RegistrationStatus, UserRole
from ....domain.services import TokenService
from ....application.use_cases.admission import ApproveAdmissionUseCase
from ....application.use_cases.registration.register_for_competition import (
    RegisterForCompetitionUseCase,
)
from ....application.use_cases.competitions.get_scoring_progress import GetScoringProgressUseCase
from ....config import settings
from ....shared.roman import arabic_to_roman
from ...schemas.admin_schemas import (
    CreateStaffRequest,
    UpdateUserRequest,
    UserListResponse,
    AdminUserResponse,
    AuditLogEntry,
    AuditLogListResponse,
    StatisticsResponse,
    AdminRegisterRequest,
    AdminRegisterResponse,
    ReplaceParticipantRequest,
    ReplaceParticipantResponse,
    AdminRegistrationItem,
    AdminRegistrationListResponse,
    AssignStaffRequest,
    CompetitionStaffItem,
    CompetitionStaffList,
    TourProgress,
    TourTimeItem,
    SetTourTimeRequest,
    ScoringProgressItem,
    ScoringProgressResponse,
    TourConfigItem,
)
from ...dependencies import require_role
from ...utils.special_import import parse_uchastniki_xlsx, derive_imported_email

router = APIRouter()


_IMPORT_HEADER_ALIASES = {
    "full_name": {
        "full_name",
        "fio",
        "фио",
        "name",
        "participant_name",
        "participant",
    },
    "email": {"email", "почта", "e-mail", "mail"},
    "institution": {
        "institution",
        "institution_name",
        "school",
        "university",
        "вуз",
        "учреждение",
        "учебное учреждение",
    },
    "institution_location": {
        "institution_location",
        "location",
        "city",
        "campus",
        "местоположение",
        "город",
        "местоположение вуза",
        "город вуза",
    },
    "is_captain": {"is_captain", "captain", "капитан", "капитан/не капитан"},
    "dob": {"dob", "birth_date", "date_of_birth", "дата рождения", "рождение"},
}


def _normalize_header(name: str) -> str:
    key = name.strip().lower()
    for canonical, aliases in _IMPORT_HEADER_ALIASES.items():
        if key in aliases:
            return canonical
    return key


def _normalize_record(raw: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for key, value in raw.items():
        normalized[_normalize_header(str(key))] = value
    return normalized


def _parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    normalized = str(value).strip().lower()
    return normalized in {"1", "true", "yes", "y", "да", "капитан"}


def _parse_dob(value: Any):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Не удалось распознать дату рождения: {value}")


def _slugify_folder_name(value: str) -> str:
    safe = re.sub(r"[^\w\-. ]+", "_", value, flags=re.UNICODE).strip()
    safe = re.sub(r"\s+", "_", safe)
    return safe[:80] or "participant"


def _derive_team_name(participant) -> str:
    """Derive team display name, appending city/location when set."""
    inst = getattr(participant, "institution", None)
    location = (getattr(participant, "institution_location", None) or "").strip()
    if inst:
        return f"{inst.name} ({location})" if location else inst.name
    return (getattr(participant, "school", None) or "").strip() or "Команда"


def _split_full_name(full_name: str) -> tuple[str, str, str]:
    parts = [part for part in re.split(r"\s+", (full_name or "").strip()) if part]
    if not parts:
        return "", "", ""
    if len(parts) == 1:
        return parts[0], "", ""
    if len(parts) == 2:
        return parts[0], parts[1], ""
    return parts[0], parts[1], " ".join(parts[2:])


def _build_badge_photo_lookup_keys(
    city: str | None,
    institution_name: str | None,
    last_name: str,
    first_name: str,
    middle_name: str,
) -> list[str]:
    fio = "_".join(part for part in [last_name, first_name, middle_name] if part).strip("_")
    if not fio:
        return []
    normalize = WordTemplateGenerator._normalize_photo_key  # type: ignore[attr-defined]
    return [
        normalize(f"{city or ''}/{institution_name or ''}/{fio}"),
        normalize(f"{institution_name or ''}/{fio}"),
        normalize(fio),
    ]


async def _load_badge_photo_index(db: AsyncSession) -> dict[str, bytes]:
    from ....infrastructure.database.models import BadgePhotoModel

    result = await db.execute(select(BadgePhotoModel.normalized_key, BadgePhotoModel.image_bytes))
    index: dict[str, bytes] = {}
    for row in result.all():
        key = (row.normalized_key or "").strip().lower()
        if not key:
            continue
        index[key] = row.image_bytes
        # Add partial-path aliases (last 1-3 segments) so that a ZIP with a root
        # folder (e.g. "photos_archive/city/inst/name") still matches lookups that
        # omit the root prefix ("city/inst/name", "inst/name", "name").
        parts = [p for p in key.split("/") if p]
        for n in range(1, min(4, len(parts))):
            alias = "/".join(parts[-n:])
            index.setdefault(alias, row.image_bytes)
    return index


def _find_badge_photo_bytes(
    photo_index: dict[str, bytes],
    city: str | None,
    institution_name: str | None,
    last_name: str,
    first_name: str,
    middle_name: str,
) -> bytes | None:
    keys = _build_badge_photo_lookup_keys(
        city=city,
        institution_name=institution_name,
        last_name=last_name,
        first_name=first_name,
        middle_name=middle_name,
    )
    for key in keys:
        if key in photo_index:
            return photo_index[key]
    return None


def _parse_import_file(file_name: str, file_bytes: bytes) -> list[dict[str, Any]]:
    lower_name = file_name.lower()

    if lower_name.endswith(".json"):
        payload = json.loads(file_bytes.decode("utf-8"))
        if isinstance(payload, dict):
            payload = payload.get("participants", [])
        if not isinstance(payload, list):
            raise ValueError("JSON должен быть массивом участников или объектом с ключом participants")
        return [_normalize_record(item) for item in payload if isinstance(item, dict)]

    if lower_name.endswith(".csv"):
        text = None
        for enc in ("utf-8-sig", "utf-8", "cp1251"):
            try:
                text = file_bytes.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            raise ValueError("Не удалось декодировать CSV. Используйте UTF-8 или CP1251")

        reader = csv.DictReader(io.StringIO(text))
        return [_normalize_record(row) for row in reader]

    if lower_name.endswith(".xlsx"):
        # First try the official «Участники» template (positional columns).
        templated = parse_uchastniki_xlsx(file_bytes)
        if templated is not None:
            return templated

        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Для импорта XLSX требуется зависимость openpyxl") from exc

        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [_normalize_header(str(h or "")) for h in rows[0]]
        records: list[dict[str, Any]] = []
        for row in rows[1:]:
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            item = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            records.append(item)
        return records

    raise ValueError("Поддерживаются только файлы .json, .csv, .xlsx")


def _extract_special_tours(competition) -> list[dict[str, Any]]:
    """Extract normalized special tours config from competition settings."""
    allowed_modes = {"individual", "individual_captains", "team"}
    settings_payload = competition.special_settings or {}
    raw_tours = settings_payload.get("tours")

    normalized: list[dict[str, Any]] = []
    if isinstance(raw_tours, list):
        for idx, item in enumerate(raw_tours, start=1):
            if not isinstance(item, dict):
                continue
            tour_number = int(item.get("tour_number") or idx)
            mode = str(item.get("mode") or "individual").strip()
            if mode not in allowed_modes:
                mode = "individual"
            task_numbers = item.get("task_numbers") or item.get("tasks") or [1]
            tasks: list[int] = []
            for t in task_numbers:
                try:
                    val = int(t)
                    if val > 0:
                        tasks.append(val)
                except Exception:  # noqa: BLE001
                    continue
            if not tasks:
                tasks = [1]
            captains_task = bool(item.get("captains_task", False))
            raw_captains_task_numbers = item.get("captains_task_numbers") or []
            captains_task_numbers: list[int] = []
            for ct in raw_captains_task_numbers:
                try:
                    val = int(ct)
                    if val > 0:
                        captains_task_numbers.append(val)
                except Exception:  # noqa: BLE001
                    continue
            normalized.append(
                {
                    "tour_number": tour_number,
                    "mode": mode,
                    "task_numbers": sorted(set(tasks)),
                    "captains_task": captains_task,
                    "captains_task_numbers": sorted(set(captains_task_numbers)) if captains_task_numbers else [1],
                }
            )

    if normalized:
        return normalized

    tours_count = int(competition.special_tours_count or 1)
    modes = competition.special_tour_modes or []
    fallback: list[dict[str, Any]] = []
    for i in range(tours_count):
        mode = modes[i] if i < len(modes) else "individual"
        if mode not in allowed_modes:
            mode = "individual"
        fallback.append(
            {
                "tour_number": i + 1,
                "mode": mode,
                "task_numbers": [1],
                "captains_task": False,
                "captains_task_numbers": [],
            }
        )
    return fallback


def _resolve_default_seat_matrix_columns(competition) -> int:
    settings_payload = competition.special_settings or {}
    raw_value = settings_payload.get("seat_matrix_columns", 3)
    try:
        columns = int(raw_value)
    except (TypeError, ValueError):
        columns = 3
    return max(columns, 1)


def _resolve_room_seat_matrix_columns(competition, room_id: UUID, is_team_mode: bool) -> int:
    settings_payload = (competition.special_settings or {}) if competition else {}
    room_key = str(room_id)

    def _extract_from_map(mapping: Any) -> int | None:
        if not isinstance(mapping, dict):
            return None
        room_payload = mapping.get(room_key)
        if not isinstance(room_payload, dict):
            return None
        raw_value = room_payload.get("seat_matrix_columns")
        try:
            parsed = int(raw_value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    room_layouts = settings_payload.get("room_layouts")
    team_room_layouts = settings_payload.get("team_room_layouts")
    selected = _extract_from_map(team_room_layouts if is_team_mode else room_layouts)
    if selected is None and is_team_mode:
        selected = _extract_from_map(room_layouts)
    if selected is None:
        selected = _resolve_default_seat_matrix_columns(competition)
    return max(int(selected), 1)


def _resolve_room_team_table_merges(competition, room_id: UUID, tour_number: int | None) -> list[list[int]]:
    if not competition or tour_number is None:
        return []

    settings_payload = competition.special_settings or {}
    raw_merges = settings_payload.get("team_table_merges")
    if not isinstance(raw_merges, dict):
        return []

    tour_payload = raw_merges.get(str(tour_number))
    if tour_payload is None:
        tour_payload = raw_merges.get(tour_number)
    if not isinstance(tour_payload, dict):
        return []

    room_payload = tour_payload.get(str(room_id))
    if not isinstance(room_payload, list):
        return []

    normalized: list[list[int]] = []
    used_tables: set[int] = set()
    for group in room_payload:
        if not isinstance(group, list):
            continue
        group_values = sorted({int(v) for v in group if isinstance(v, int) and int(v) > 0})
        if len(group_values) < 2:
            continue
        if any(v in used_tables for v in group_values):
            continue
        used_tables.update(group_values)
        normalized.append(group_values)
    return normalized


def _resolve_special_tour_context(competition, tour_number: int | None) -> dict[str, Any] | None:
    if not competition or not getattr(competition, "is_special", False):
        return None

    tours = _extract_special_tours(competition)
    if not tours:
        return None

    if tour_number is None:
        selected = tours[0]
    else:
        selected = next((item for item in tours if int(item.get("tour_number", 0)) == tour_number), None)
        if not selected:
            raise HTTPException(status_code=400, detail="Указан несуществующий номер тура")

    mode = str(selected.get("mode") or "individual").strip()
    return {
        "tour_number": int(selected.get("tour_number") or 1),
        "mode": mode,
        "is_team_mode": mode == "team",
    }


def _resolve_room_seats_per_table(competition, room_id: UUID, is_team_mode: bool) -> int:
    settings_payload = (competition.special_settings or {}) if competition else {}
    room_key = str(room_id)

    def _extract_from_map(mapping: Any) -> int | None:
        if not isinstance(mapping, dict):
            return None
        room_payload = mapping.get(room_key)
        if not isinstance(room_payload, dict):
            return None
        raw_value = room_payload.get("seats_per_table")
        try:
            parsed = int(raw_value)
        except (TypeError, ValueError):
            return None
        return parsed if parsed > 0 else None

    room_layouts = settings_payload.get("room_layouts")
    team_room_layouts = settings_payload.get("team_room_layouts")
    selected = _extract_from_map(team_room_layouts if is_team_mode else room_layouts)
    if selected is None and is_team_mode:
        selected = _extract_from_map(room_layouts)

    if selected is None:
        raw_default = (
            settings_payload.get("team_default_seats_per_table")
            if is_team_mode
            else settings_payload.get("default_seats_per_table")
        )
        try:
            selected = int(raw_default)
        except (TypeError, ValueError):
            selected = 1

    return max(int(selected), 1)


def _build_room_tables(
    room_capacity: int,
    seats_by_number: dict[int, dict[str, Any]],
    seats_per_table: int,
) -> list[dict[str, Any]]:
    tables: list[dict[str, Any]] = []
    tables_count = (room_capacity + seats_per_table - 1) // seats_per_table
    for table_number in range(1, tables_count + 1):
        table_seats: list[dict[str, Any]] = []
        base = (table_number - 1) * seats_per_table
        for seat_at_table in range(1, seats_per_table + 1):
            seat_number = base + seat_at_table
            if seat_number > room_capacity:
                break
            seat_data = seats_by_number.get(seat_number)
            table_seats.append(
                {
                    "seat_number": seat_number,
                    "seat_at_table": seat_at_table,
                    "table_number": table_number,
                    "occupied": seat_data is not None,
                    "variant_number": seat_data["variant_number"] if seat_data else None,
                    "participant_name": seat_data["participant_name"] if seat_data else None,
                    "institution_id": seat_data["institution_id"] if seat_data else None,
                    "institution_name": seat_data["institution_name"] if seat_data else None,
                    "institution_location": seat_data["institution_location"] if seat_data else None,
                    "is_captain": seat_data["is_captain"] if seat_data else False,
                }
            )
        if table_seats:
            tables.append(
                {
                    "table_number": table_number,
                    "occupied": any(seat["occupied"] for seat in table_seats),
                    "seats": table_seats,
                }
            )
    return tables


def _annotate_tables_with_merges(room_tables: list[dict[str, Any]], merge_groups: list[list[int]]) -> list[list[int]]:
    if not room_tables:
        return []

    table_numbers = {int(table.get("table_number", 0)) for table in room_tables}
    valid_groups: list[list[int]] = []
    table_to_group: dict[int, int] = {}

    for group in merge_groups:
        valid_group = sorted(table for table in group if table in table_numbers)
        if len(valid_group) < 2:
            continue
        group_id = len(valid_groups) + 1
        valid_groups.append(valid_group)
        for table_number in valid_group:
            table_to_group[table_number] = group_id

    for table in room_tables:
        table_number = int(table.get("table_number", 0))
        group_id = table_to_group.get(table_number)
        table["merged_group"] = group_id
        table["merged_with"] = (
            [n for n in valid_groups[group_id - 1] if n != table_number]
            if group_id
            else []
        )

    return valid_groups


def _project_team_seating_for_merges(
    room_tables: list[dict[str, Any]],
    merge_groups: list[list[int]],
) -> None:
    """Rebuild room table occupancy for team mode so merged groups hold one team when possible.

    This is a view-layer projection for team seating plan rendering (does not persist to DB).
    """
    if not room_tables:
        return

    table_by_number: dict[int, dict[str, Any]] = {
        int(table.get("table_number", 0)): table
        for table in room_tables
        if int(table.get("table_number", 0)) > 0
    }
    if not table_by_number:
        return

    valid_groups: list[list[int]] = []
    merged_table_numbers: set[int] = set()
    for group in merge_groups:
        numbers = sorted({int(number) for number in group if int(number) in table_by_number})
        if len(numbers) < 2:
            continue
        valid_groups.append(numbers)
        merged_table_numbers.update(numbers)

    # Collect currently seated participants and group by institution (team).
    team_buckets: dict[str, list[dict[str, Any]]] = {}
    for table in room_tables:
        for seat in table.get("seats", []):
            if not seat.get("occupied"):
                continue
            institution_id = seat.get("institution_id")
            institution_name = (seat.get("institution_name") or "").strip()
            institution_location_val = (seat.get("institution_location") or "").strip().lower()
            if institution_id:
                team_key = f"id:{institution_id}:{institution_location_val}" if institution_location_val else f"id:{institution_id}"
            elif institution_name:
                team_key = f"name:{institution_name.lower()}"
            else:
                team_key = f"solo:{seat.get('participant_name') or seat.get('seat_number')}"

            team_buckets.setdefault(team_key, []).append(
                {
                    "occupied": True,
                    "variant_number": seat.get("variant_number"),
                    "participant_name": seat.get("participant_name"),
                    "institution_id": seat.get("institution_id"),
                    "institution_name": seat.get("institution_name"),
                    "institution_location": seat.get("institution_location"),
                    "is_captain": bool(seat.get("is_captain")),
                }
            )

    def _active_team_keys() -> list[str]:
        return sorted(
            (key for key, bucket in team_buckets.items() if bucket),
            key=lambda key: (-len(team_buckets[key]), key),
        )

    def _take_from_team(team_key: str, count: int) -> list[dict[str, Any]]:
        if count <= 0:
            return []
        bucket = team_buckets.get(team_key) or []
        taken = bucket[:count]
        team_buckets[team_key] = bucket[count:]
        return taken

    def _empty_seat_payload() -> dict[str, Any]:
        return {
            "occupied": False,
            "variant_number": None,
            "participant_name": None,
            "institution_id": None,
            "institution_name": None,
            "institution_location": None,
            "is_captain": False,
        }

    # Reset all seats first.
    for table in room_tables:
        for seat in table.get("seats", []):
            seat.update(_empty_seat_payload())

    def _slots_for_table_number(table_number: int) -> list[dict[str, Any]]:
        table = table_by_number.get(table_number)
        if not table:
            return []
        return sorted(
            list(table.get("seats", [])),
            key=lambda seat: (int(seat.get("seat_at_table") or 0), int(seat.get("seat_number") or 0)),
        )

    def _assign_to_slots(slots: list[dict[str, Any]], payloads: list[dict[str, Any]]) -> list[dict[str, Any]]:
        for slot, payload in zip(slots, payloads):
            slot.update(payload)
        return slots[len(payloads):]

    merged_overflow_slots: list[dict[str, Any]] = []

    # Pass 1: each merged block gets one team whenever possible.
    for group in valid_groups:
        slots: list[dict[str, Any]] = []
        for table_number in group:
            slots.extend(_slots_for_table_number(table_number))
        if not slots:
            continue
        keys = _active_team_keys()
        if not keys:
            merged_overflow_slots.extend(slots)
            continue
        chosen_team = keys[0]
        assigned = _take_from_team(chosen_team, len(slots))
        remaining = _assign_to_slots(slots, assigned)
        merged_overflow_slots.extend(remaining)

    # Pass 2: fill non-merged tables from remaining teams.
    non_merged_numbers = sorted(number for number in table_by_number if number not in merged_table_numbers)
    for table_number in non_merged_numbers:
        slots = _slots_for_table_number(table_number)
        slot_index = 0
        while slot_index < len(slots):
            keys = _active_team_keys()
            if not keys:
                break
            chosen_team = keys[0]
            take_count = min(len(slots) - slot_index, len(team_buckets[chosen_team]))
            assigned = _take_from_team(chosen_team, take_count)
            remaining = _assign_to_slots(slots[slot_index : slot_index + take_count], assigned)
            slot_index += take_count - len(remaining)

    # Pass 3: if participants still remain, place them into free slots inside merged blocks.
    for slot in merged_overflow_slots:
        keys = _active_team_keys()
        if not keys:
            break
        chosen_team = keys[0]
        assigned = _take_from_team(chosen_team, 1)
        if not assigned:
            continue
        slot.update(assigned[0])

    for table in room_tables:
        table["occupied"] = any(bool(seat.get("occupied")) for seat in table.get("seats", []))


async def _ensure_seat_assignments_for_competition(competition, db: AsyncSession) -> int:
    """Auto-assign seats for active registrations missing seat assignment.

    Used by seating-plan view to keep schema non-empty without heavy per-registration queries.
    Rule for `individual` / `individual_captains`: participants from the same
    institution are not placed in adjacent seats (left/right/front/back/diagonal)
    when there is any valid non-conflicting seat.
    """
    from ....infrastructure.database.models import (
        ParticipantModel,
        RegistrationModel,
        RoomModel,
        SeatAssignmentModel,
    )

    rooms_result = await db.execute(
        select(RoomModel)
        .where(RoomModel.competition_id == competition.id)
        .order_by(RoomModel.name.asc())
    )
    rooms = rooms_result.scalars().all()
    if not rooms:
        return 0

    room_ids = [room.id for room in rooms]
    rooms_by_id = {room.id: room for room in rooms}

    existing_result = await db.execute(
        select(
            SeatAssignmentModel.room_id,
            SeatAssignmentModel.seat_number,
            ParticipantModel.institution_id,
        )
        .join(RegistrationModel, RegistrationModel.id == SeatAssignmentModel.registration_id)
        .join(ParticipantModel, ParticipantModel.id == RegistrationModel.participant_id)
        .where(SeatAssignmentModel.room_id.in_(room_ids))
    )
    taken_by_room: dict[UUID, set[int]] = {room.id: set() for room in rooms}
    institution_seats_by_room: dict[UUID, dict[UUID, list[int]]] = {room.id: {} for room in rooms}
    for row in existing_result:
        taken_by_room[row.room_id].add(int(row.seat_number))
        if row.institution_id:
            room_map = institution_seats_by_room[row.room_id]
            room_map.setdefault(row.institution_id, []).append(int(row.seat_number))

    missing_result = await db.execute(
        select(
            RegistrationModel.id.label("registration_id"),
            ParticipantModel.institution_id.label("institution_id"),
            ParticipantModel.is_captain.label("is_captain"),
        )
        .join(ParticipantModel, ParticipantModel.id == RegistrationModel.participant_id)
        .outerjoin(SeatAssignmentModel, SeatAssignmentModel.registration_id == RegistrationModel.id)
        .where(RegistrationModel.competition_id == competition.id)
        .where(RegistrationModel.status != RegistrationStatus.CANCELLED)
        .where(SeatAssignmentModel.id.is_(None))
        .order_by(RegistrationModel.created_at.asc())
    )
    missing_rows = missing_result.all()
    if not missing_rows:
        return 0

    special_context = _resolve_special_tour_context(competition, tour_number=None)
    tour_mode = str(special_context.get("mode")) if special_context else ""
    is_captains_mode = tour_mode == "individual_captains"
    apply_institution_distance_rule = tour_mode in {"individual", "individual_captains"}
    is_team_mode = bool(special_context and special_context.get("is_team_mode"))
    settings_payload = competition.special_settings or {}
    captains_room_id: UUID | None = None
    raw_captains_room_id = settings_payload.get("captains_room_id")
    if is_captains_mode and raw_captains_room_id:
        try:
            parsed = UUID(str(raw_captains_room_id))
            if parsed in rooms_by_id:
                captains_room_id = parsed
        except (TypeError, ValueError):
            captains_room_id = None

    room_priority_all = [room.id for room in rooms]
    room_columns = {
        room.id: _resolve_room_seat_matrix_columns(
            competition=competition,
            room_id=room.id,
            is_team_mode=is_team_mode,
        )
        for room in rooms
    }
    room_seats_per_table = {
        room.id: _resolve_room_seats_per_table(
            competition=competition,
            room_id=room.id,
            is_team_mode=is_team_mode,
        )
        for room in rooms
    }

    def seat_grid_position(room_id: UUID, seat_number: int) -> tuple[int, int]:
        cols = max(int(room_columns.get(room_id, 3)), 1)
        row = (seat_number - 1) // cols
        col = (seat_number - 1) % cols
        return row, col

    def same_institution_neighborhood_conflicts(room_id: UUID, seat_number: int, institution_id: UUID | None) -> int:
        if not institution_id:
            return 0
        target_row, target_col = seat_grid_position(room_id, seat_number)
        conflicts = 0
        for occupied_seat in institution_seats_by_room[room_id].get(institution_id, []):
            occ_row, occ_col = seat_grid_position(room_id, occupied_seat)
            if abs(occ_row - target_row) <= 1 and abs(occ_col - target_col) <= 1:
                conflicts += 1
        return conflicts

    def same_institution_table_conflicts(room_id: UUID, seat_number: int, institution_id: UUID | None) -> int:
        if not institution_id:
            return 0
        seats_per_table = max(int(room_seats_per_table.get(room_id, 1)), 1)
        target_table = (seat_number - 1) // seats_per_table
        conflicts = 0
        for occupied_seat in institution_seats_by_room[room_id].get(institution_id, []):
            occupied_table = (occupied_seat - 1) // seats_per_table
            if occupied_table == target_table:
                conflicts += 1
        return conflicts

    def candidate_room_order(is_captain: bool) -> list[UUID]:
        if not is_captains_mode or captains_room_id is None:
            return room_priority_all
        if is_captain:
            return [captains_room_id] + [rid for rid in room_priority_all if rid != captains_room_id]
        return [rid for rid in room_priority_all if rid != captains_room_id] + [captains_room_id]

    seat_repo = SeatAssignmentRepositoryImpl(db)
    variants_count = max(int(getattr(competition, "variants_count", 1) or 1), 1)
    assigned_count = 0

    for missing in missing_rows:
        best_strict: tuple[tuple[int, int, int, int], UUID, int] | None = None
        best_relaxed: tuple[tuple[int, int, int, int], UUID, int] | None = None

        for room_priority_index, room_id in enumerate(candidate_room_order(bool(missing.is_captain))):
            room = rooms_by_id[room_id]
            taken = taken_by_room[room_id]
            for seat_number in range(1, room.capacity + 1):
                if seat_number in taken:
                    continue

                neighborhood_conflicts = (
                    same_institution_neighborhood_conflicts(room_id, seat_number, missing.institution_id)
                    if apply_institution_distance_rule
                    else 0
                )
                table_conflicts = (
                    same_institution_table_conflicts(room_id, seat_number, missing.institution_id)
                    if apply_institution_distance_rule
                    else 0
                )
                same_inst_in_room = (
                    len(institution_seats_by_room[room_id].get(missing.institution_id, []))
                    if missing.institution_id
                    else 0
                )

                score = (same_inst_in_room, room_priority_index, len(taken), seat_number)
                if neighborhood_conflicts == 0 and table_conflicts == 0:
                    candidate = (score, room_id, seat_number)
                    if best_strict is None or candidate[0] < best_strict[0]:
                        best_strict = candidate
                else:
                    relaxed_score = (
                        neighborhood_conflicts + table_conflicts,
                        same_inst_in_room,
                        room_priority_index,
                        seat_number,
                    )
                    candidate = (relaxed_score, room_id, seat_number)
                    if best_relaxed is None or candidate[0] < best_relaxed[0]:
                        best_relaxed = candidate

        selected = best_strict or best_relaxed
        if selected is None:
            continue
        _, chosen_room_id, chosen_seat_number = selected

        variant_number = (chosen_seat_number % variants_count) + 1
        await seat_repo.create(
            SeatAssignment(
                registration_id=missing.registration_id,
                room_id=chosen_room_id,
                seat_number=chosen_seat_number,
                variant_number=variant_number,
            )
        )
        taken_by_room[chosen_room_id].add(chosen_seat_number)
        if missing.institution_id:
            institution_seats_by_room[chosen_room_id].setdefault(missing.institution_id, []).append(chosen_seat_number)
        assigned_count += 1

    return assigned_count


# --- User Management ---

@router.get("/users", response_model=UserListResponse)
async def list_users(
    skip: int = 0,
    limit: int = 50,
    role: Optional[UserRole] = None,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """List all users with optional role filter."""
    user_repo = UserRepositoryImpl(db)

    if role:
        users = await user_repo.get_by_role(role, skip=skip, limit=limit)
    else:
        users = await user_repo.get_all(skip=skip, limit=limit)

    items = [
        AdminUserResponse(
            id=u.id,
            email=u.email,
            role=u.role,
            is_active=u.is_active,
            created_at=u.created_at,
        )
        for u in users
    ]
    return UserListResponse(items=items, total=len(items))


@router.post("/users", response_model=AdminUserResponse, status_code=status.HTTP_201_CREATED)
async def create_staff_user(
    body: CreateStaffRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Create a user account (participant / admitter / scanner / admin)."""
    user_repo = UserRepositoryImpl(db)

    if await user_repo.exists_by_email(body.email):
        raise HTTPException(status_code=400, detail="Email уже используется")

    # Validate participant-specific fields
    if body.role == UserRole.PARTICIPANT:
        if not body.full_name or len(body.full_name.strip()) < 2:
            raise HTTPException(status_code=400, detail="ФИО обязательно для участников (минимум 2 символа)")
        if not body.school or len(body.school.strip()) < 2:
            raise HTTPException(status_code=400, detail="Учебное учреждение обязательно для участников (минимум 2 символа)")

    from uuid import uuid4

    user = User(
        id=uuid4(),
        email=body.email,
        password_hash=hash_password(body.password),
        role=body.role,
    )
    user = await user_repo.create(user)

    # Create participant profile if role is participant
    if body.role == UserRole.PARTICIPANT:
        from ....domain.entities import Participant
        participant_repo = ParticipantRepositoryImpl(db)

        participant = Participant(
            id=uuid4(),
            user_id=user.id,
            full_name=body.full_name,
            school=body.school,
            grade=body.grade,
            institution_id=body.institution_id,
            institution_location=body.institution_location,
            is_captain=body.is_captain,
            dob=body.dob,
        )
        await participant_repo.create(participant)

    return AdminUserResponse(
        id=user.id,
        email=user.email,
        role=user.role,
        is_active=user.is_active,
        created_at=user.created_at,
    )


@router.put("/users/{user_id}", response_model=AdminUserResponse)
async def update_user(
    user_id: UUID,
    body: UpdateUserRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Update user attributes (active status, role)."""
    user_repo = UserRepositoryImpl(db)
    user = await user_repo.get_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    if body.is_active is not None:
        if body.is_active:
            user.activate()
        else:
            user.deactivate()

    if body.role is not None:
        user.change_role(body.role)

    await user_repo.update(user)

    return AdminUserResponse(
        id=user.id,
        email=user.email,
        role=user.role,
        is_active=user.is_active,
        created_at=user.created_at,
    )


@router.delete("/users/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
async def deactivate_user(
    user_id: UUID,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Deactivate a user (soft delete)."""
    user_repo = UserRepositoryImpl(db)
    user = await user_repo.get_by_id(user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")

    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="Нельзя деактивировать себя")

    user.deactivate()
    await user_repo.update(user)


# --- Participants ---

@router.get("/participants")
async def list_participants(
    skip: int = 0,
    limit: int = 1000,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """List all participants (id, full_name, school) for admin registration."""
    from ....infrastructure.database.models import ParticipantModel

    stmt = (
        select(ParticipantModel)
        .order_by(ParticipantModel.full_name)
        .offset(skip)
        .limit(limit)
    )
    result = await db.execute(stmt)
    participants = result.scalars().all()

    return {
        "participants": [
            {
                "id": str(p.id),
                "user_id": str(p.user_id),
                "full_name": p.full_name,
                "school": p.school,
                "institution_location": p.institution_location,
                "is_captain": p.is_captain,
            }
            for p in participants
        ]
    }


@router.patch("/participants/{participant_id}")
async def update_participant_fields(
    participant_id: UUID,
    institution_location: Optional[str] = None,
    institution_id: Optional[UUID] = None,
    clear_institution: bool = False,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Update institution_location and/or institution_id for a participant."""
    from ....infrastructure.database.models import ParticipantModel

    result = await db.execute(select(ParticipantModel).where(ParticipantModel.id == participant_id))
    model = result.scalar_one_or_none()
    if not model:
        raise HTTPException(status_code=404, detail="Участник не найден")

    if institution_location is not None:
        model.institution_location = institution_location or None
    if institution_id is not None:
        model.institution_id = institution_id
    if clear_institution:
        model.institution_id = None

    await db.flush()
    return {
        "id": str(model.id),
        "full_name": model.full_name,
        "institution_location": model.institution_location,
        "institution_id": str(model.institution_id) if model.institution_id else None,
    }


@router.post("/participants/{participant_id}/badge-photo")
async def upload_participant_badge_photo(
    participant_id: UUID,
    file: UploadFile = File(...),
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Upload a badge photo for a specific participant. Overwrites any existing photo for them."""
    from ....infrastructure.database.models import ParticipantModel, BadgePhotoModel

    result = await db.execute(
        select(ParticipantModel)
        .where(ParticipantModel.id == participant_id)
        .options(selectinload(ParticipantModel.institution))
    )
    participant = result.scalar_one_or_none()
    if not participant:
        raise HTTPException(status_code=404, detail="Участник не найден")

    image_bytes = await file.read()
    if not image_bytes:
        raise HTTPException(status_code=400, detail="Файл пустой")

    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    allowed = {"jpg", "jpeg", "png", "webp"}
    if ext not in allowed:
        raise HTTPException(status_code=400, detail="Допустимые форматы: JPG, PNG, WEBP")
    content_type = f"image/{ext}" if ext != "jpg" else "image/jpeg"

    last_name, first_name, middle_name = _split_full_name(participant.full_name)
    fio = "_".join(p for p in [last_name, first_name, middle_name] if p)
    institution_name = participant.institution.name if participant.institution else ""
    city = participant.institution_location or ""

    normalize = WordTemplateGenerator._normalize_photo_key  # type: ignore[attr-defined]
    normalized_key = normalize(f"{city}/{institution_name}/{fio}")
    original_path = f"{city}/{institution_name}/{fio}.{ext}"

    # Upsert: delete existing record with same key, then insert
    await db.execute(delete(BadgePhotoModel).where(BadgePhotoModel.normalized_key == normalized_key))
    db.add(BadgePhotoModel(
        normalized_key=normalized_key,
        original_path=original_path,
        content_type=content_type,
        image_bytes=image_bytes,
    ))
    await db.flush()
    return {"ok": True, "normalized_key": normalized_key}


# --- Audit Log ---

@router.get("/audit-log", response_model=AuditLogListResponse)
async def list_audit_log(
    skip: int = 0,
    limit: int = 50,
    entity_type: Optional[str] = None,
    entity_id: Optional[UUID] = None,
    user_id: Optional[UUID] = None,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """List audit log entries with optional filters."""
    audit_repo = AuditLogRepositoryImpl(db)

    if entity_type and entity_id:
        logs = await audit_repo.get_by_entity(entity_type, entity_id, skip=skip, limit=limit)
    elif user_id:
        logs = await audit_repo.get_by_user(user_id, skip=skip, limit=limit)
    else:
        logs = await audit_repo.get_all(skip=skip, limit=limit)

    items = [
        AuditLogEntry(
            id=log.id,
            entity_type=log.entity_type,
            entity_id=log.entity_id,
            action=log.action,
            user_id=log.user_id,
            ip_address=log.ip_address,
            details=log.details,
            timestamp=log.timestamp,
        )
        for log in logs
    ]
    return AuditLogListResponse(items=items, total=len(items))


# --- Statistics ---

@router.get("/statistics", response_model=StatisticsResponse)
async def get_statistics(
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Get system statistics for admin dashboard."""
    from sqlalchemy import select, func
    from ....infrastructure.database.models import (
        UserModel,
        CompetitionModel,
        ScanModel,
        RegistrationModel,
        ParticipantModel,
    )

    # Count users
    result = await db.execute(select(func.count()).select_from(UserModel))
    total_users = result.scalar() or 0

    # Count competitions
    result = await db.execute(select(func.count()).select_from(CompetitionModel))
    total_competitions = result.scalar() or 0

    # Count scans
    result = await db.execute(select(func.count()).select_from(ScanModel))
    total_scans = result.scalar() or 0

    # Count registrations
    result = await db.execute(select(func.count()).select_from(RegistrationModel))
    total_registrations = result.scalar() or 0

    # Count participants
    result = await db.execute(select(func.count()).select_from(ParticipantModel))
    total_participants = result.scalar() or 0

    return StatisticsResponse(
        total_competitions=total_competitions,
        total_users=total_users,
        total_scans=total_scans,
        total_registrations=total_registrations,
        total_participants=total_participants,
    )


# --- Registration Management ---

@router.post("/registrations", response_model=AdminRegisterResponse, status_code=status.HTTP_201_CREATED)
async def admin_register_participant(
    body: AdminRegisterRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Admin registers a participant for a competition (bypasses status check)."""
    registration_repo = RegistrationRepositoryImpl(db)
    competition_repo = CompetitionRepositoryImpl(db)
    participant_repo = ParticipantRepositoryImpl(db)
    entry_token_repo = EntryTokenRepositoryImpl(db)
    token_service = TokenService(settings.hmac_secret_key)

    use_case = RegisterForCompetitionUseCase(
        registration_repository=registration_repo,
        competition_repository=competition_repo,
        participant_repository=participant_repo,
        entry_token_repository=entry_token_repo,
        token_service=token_service,
    )

    try:
        result = await use_case.execute(
            participant_id=body.participant_id,
            competition_id=body.competition_id,
            skip_status_check=True,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    return AdminRegisterResponse(
        registration_id=result.registration_id,
        entry_token=result.entry_token,
    )


@router.get("/registrations/{competition_id}", response_model=AdminRegistrationListResponse)
async def list_competition_registrations(
    competition_id: UUID,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """List all registrations for a competition with participant details."""
    from ....infrastructure.database.models import (
        RegistrationModel,
        ParticipantModel,
        EntryTokenModel,
        InstitutionModel,
    )

    stmt = (
        select(RegistrationModel)
        .where(RegistrationModel.competition_id == competition_id)
        .options(
            selectinload(RegistrationModel.entry_token),
            selectinload(RegistrationModel.participant).selectinload(ParticipantModel.institution),
        )
        .order_by(RegistrationModel.created_at)
    )
    result = await db.execute(stmt)
    registrations = result.scalars().all()

    # Bulk-load seat assignments for all registrations in one query
    reg_ids = [reg.id for reg in registrations]
    seat_map: dict[UUID, tuple[str | None, int | None, int | None]] = {}
    if reg_ids:
        seat_stmt = (
            select(SeatAssignmentModel, RoomModel.name)
            .join(RoomModel, SeatAssignmentModel.room_id == RoomModel.id)
            .where(SeatAssignmentModel.registration_id.in_(reg_ids))
        )
        seat_result = await db.execute(seat_stmt)
        for seat_row, room_name in seat_result.all():
            seat_map[seat_row.registration_id] = (room_name, seat_row.seat_number, seat_row.variant_number)

    items = []
    for reg in registrations:
        participant = reg.participant
        institution_name = None
        if participant and participant.institution:
            institution_name = participant.institution.name

        entry_token_raw = None
        if reg.entry_token:
            entry_token_raw = reg.entry_token.raw_token

        seat_info = seat_map.get(reg.id)
        items.append(
            AdminRegistrationItem(
                registration_id=reg.id,
                participant_id=reg.participant_id,
                participant_name=participant.full_name if participant else "—",
                participant_school=participant.school if participant else "—",
                participant_institution_location=participant.institution_location if participant else None,
                participant_is_captain=participant.is_captain if participant else False,
                institution_name=institution_name,
                entry_token=entry_token_raw,
                status=reg.status.value,
                seat_room_name=seat_info[0] if seat_info else None,
                seat_number=seat_info[1] if seat_info else None,
                variant_number=seat_info[2] if seat_info else None,
            )
        )

    return AdminRegistrationListResponse(items=items, total=len(items))


@router.delete("/registrations/{registration_id}", status_code=status.HTTP_200_OK)
async def delete_registration(
    registration_id: UUID,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Delete (remove) a participant registration from a competition."""
    registration_repo = RegistrationRepositoryImpl(db)
    reg = await registration_repo.get_by_id(registration_id)
    if not reg:
        raise HTTPException(status_code=404, detail="Регистрация не найдена")
    await registration_repo.delete(registration_id)
    return {"ok": True}


@router.post("/registrations/{registration_id}/replace", response_model=ReplaceParticipantResponse)
async def replace_registration_participant(
    registration_id: UUID,
    body: ReplaceParticipantRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Replace the participant in a registration, transferring the seat assignment to the new participant."""
    registration_repo = RegistrationRepositoryImpl(db)
    competition_repo = CompetitionRepositoryImpl(db)
    participant_repo = ParticipantRepositoryImpl(db)
    entry_token_repo = EntryTokenRepositoryImpl(db)
    token_service = TokenService(settings.hmac_secret_key)

    old_reg = await registration_repo.get_by_id(registration_id)
    if not old_reg:
        raise HTTPException(status_code=404, detail="Регистрация не найдена")

    # Capture seat assignment before deletion
    seat_result = await db.execute(
        select(SeatAssignmentModel, RoomModel.name)
        .join(RoomModel, SeatAssignmentModel.room_id == RoomModel.id)
        .where(SeatAssignmentModel.registration_id == registration_id)
    )
    seat_row = seat_result.one_or_none()
    old_seat_model = seat_row[0] if seat_row else None
    old_room_name = seat_row[1] if seat_row else None

    warning = None
    if old_reg.status.value == "completed":
        warning = "Старый участник уже был допущен — его попытка удалена"

    # Create new registration for the replacement participant
    use_case = RegisterForCompetitionUseCase(
        registration_repository=registration_repo,
        competition_repository=competition_repo,
        participant_repository=participant_repo,
        entry_token_repository=entry_token_repo,
        token_service=token_service,
    )
    try:
        new_reg_result = await use_case.execute(
            participant_id=body.new_participant_id,
            competition_id=old_reg.competition_id,
            skip_status_check=True,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    new_registration_id = new_reg_result.registration_id

    # Delete old registration (CASCADE removes its seat_assignment, entry_token, attempts, scans)
    await registration_repo.delete(registration_id)

    # Re-assign same seat to the new registration
    if old_seat_model is not None:
        await db.execute(
            insert(SeatAssignmentModel).values(
                id=uuid4(),
                registration_id=new_registration_id,
                room_id=old_seat_model.room_id,
                seat_number=old_seat_model.seat_number,
                variant_number=old_seat_model.variant_number,
                created_at=datetime.utcnow(),
            )
        )
        await db.flush()

    return ReplaceParticipantResponse(
        new_registration_id=new_registration_id,
        entry_token=new_reg_result.entry_token,
        seat_transferred=old_seat_model is not None,
        room_name=old_room_name if old_seat_model else None,
        seat_number=old_seat_model.seat_number if old_seat_model else None,
        variant_number=old_seat_model.variant_number if old_seat_model else None,
        warning=warning,
    )


@router.post("/registrations/{registration_id}/admit-and-download")
async def admit_and_download_single(
    registration_id: UUID,
    request: Request,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Admit a single participant (if pending) and download their answer sheets + badge as ZIP."""
    from ....infrastructure.database.models import (
        RegistrationModel,
        ParticipantModel,
    )
    from io import BytesIO

    # Load registration with all relationships
    stmt = (
        select(RegistrationModel)
        .where(RegistrationModel.id == registration_id)
        .options(
            selectinload(RegistrationModel.participant).selectinload(ParticipantModel.institution),
            selectinload(RegistrationModel.entry_token),
            selectinload(RegistrationModel.attempts),
        )
    )
    result = await db.execute(stmt)
    reg = result.scalar_one_or_none()
    if not reg:
        raise HTTPException(status_code=404, detail="Регистрация не найдена")

    participant = reg.participant
    if not participant:
        raise HTTPException(status_code=400, detail="Участник не найден")

    competition_repo = CompetitionRepositoryImpl(db)
    competition = await competition_repo.get_by_id(reg.competition_id)
    if not competition:
        raise HTTPException(status_code=404, detail="Олимпиада не найдена")

    attempt_repo = AttemptRepositoryImpl(db)

    # Try to get existing attempt first (bypasses caching issues)
    attempt = await attempt_repo.get_by_registration(registration_id)

    if attempt is None and reg.status.value == "pending":
        if not reg.entry_token or not reg.entry_token.raw_token:
            raise HTTPException(status_code=400, detail="У регистрации отсутствует entry token")

        # Admin bypass: extend expired token so ApproveAdmissionUseCase can proceed
        if datetime.utcnow() > reg.entry_token.expires_at:
            from ....infrastructure.database.models import EntryTokenModel
            from datetime import timedelta
            await db.execute(
                update(EntryTokenModel)
                .where(EntryTokenModel.id == reg.entry_token.id)
                .values(expires_at=datetime.utcnow() + timedelta(hours=2))
            )
            await db.flush()

        approve_uc = ApproveAdmissionUseCase(
            token_service=TokenService(settings.hmac_secret_key),
            entry_token_repository=EntryTokenRepositoryImpl(db),
            registration_repository=RegistrationRepositoryImpl(db),
            competition_repository=competition_repo,
            attempt_repository=attempt_repo,
            audit_log_repository=AuditLogRepositoryImpl(db),
            answer_sheet_repository=AnswerSheetRepositoryImpl(db),
            storage=MinIOStorage(),
            sheet_generator=SheetGenerator(),
            room_repository=RoomRepositoryImpl(db),
            seat_assignment_repository=SeatAssignmentRepositoryImpl(db),
            participant_repository=ParticipantRepositoryImpl(db),
        )
        try:
            await approve_uc.execute(
                registration_id=reg.id,
                raw_entry_token=reg.entry_token.raw_token,
                admitter_user_id=current_user.id,
                ip_address=request.client.host if request.client else None,
            )
        except Exception as exc:
            raise HTTPException(status_code=400, detail=f"Ошибка допуска: {exc}")

        # Query attempt directly — avoids SQLAlchemy identity-map stale cache
        attempt = await attempt_repo.get_by_registration(registration_id)

    if attempt is None:
        raise HTTPException(
            status_code=400,
            detail="У участника нет попытки. Зарегистрируйте его через эту панель (кнопка ↓) или через стойку регистрации.",
        )

    folder = _slugify_folder_name(f"{participant.full_name}")

    # Load badge template for this competition (JSON-based, from visual editor)
    badge_tmpl_result = await db.execute(
        select(BadgeTemplateModel).where(BadgeTemplateModel.competition_id == reg.competition_id)
    )
    badge_template = badge_tmpl_result.scalar_one_or_none()

    zip_buffer = BytesIO()
    word_generator = WordTemplateGenerator()
    added_files = 0
    gen_errors: list[str] = []

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        # --- Answer sheets (special olympiad) ---
        if competition.is_special:
            tours = _extract_special_tours(competition)
            has_team_tours = any(str(t["mode"]) == "team" for t in tours)
            mode_labels = {
                "individual": "Индивидуальный",
                "individual_captains": "Индивидуальный (капитаны)",
                "team": "Командный",
            }
            is_captain = getattr(participant, "is_captain", False)
            institution_name_for_team = _derive_team_name(participant)
            team_folder = f"Командный зачет/{_slugify_folder_name(institution_name_for_team)}"

            for tour in tours:
                tour_number = int(tour["tour_number"])
                mode = str(tour["mode"])
                mode_label = mode_labels.get(mode, mode)
                task_numbers = tour["task_numbers"]
                is_team_tour = mode == "team"

                # Team tours: only generate sheets for captains (one set per institution)
                if is_team_tour and not is_captain:
                    continue

                # Determine folder root based on tour type and whether split is needed
                if has_team_tours:
                    tour_folder_root = team_folder if is_team_tour else f"Личный зачет/{folder}"
                else:
                    tour_folder_root = folder

                cover_qr_payload = f"attempt:{attempt.id}:tour:{tour_number}:cover"
                try:
                    cover_docx = word_generator.generate_a3_cover(
                        qr_payload=cover_qr_payload,
                        tour_number=tour_number,
                        mode=mode_label,
                    )
                    zf.writestr(f"{tour_folder_root}/tour_{tour_number}/A3_tour_{tour_number}.docx", cover_docx)
                    added_files += 1
                except Exception as exc:
                    gen_errors.append(f"A3 тур {tour_number}: {exc}")

                for task_number in task_numbers:
                    task_qr_payload = f"attempt:{attempt.id}:tour:{tour_number}:task:{task_number}"
                    try:
                        task_docx = word_generator.generate_answer_blank(
                            qr_payload=task_qr_payload,
                            tour_number=tour_number,
                            task_number=int(task_number),
                            mode=mode_label,
                        )
                        task_folder = f"{tour_folder_root}/tour_{tour_number}/task_{task_number}"
                        zf.writestr(f"{task_folder}/task_{task_number}.docx", task_docx)
                        added_files += 1
                        for extra_i in range(1, 6):
                            extra_docx = word_generator.generate_answer_blank(
                                qr_payload=task_qr_payload,
                                tour_number=tour_number,
                                task_number=int(task_number),
                                mode=mode_label,
                                tour_task=f"{tour_number}/{task_number}/{extra_i}",
                            )
                            zf.writestr(
                                f"{task_folder}/дополнительные бланки/extra_{extra_i}.docx",
                                extra_docx,
                            )
                            added_files += 1
                    except Exception as exc:
                        gen_errors.append(f"Задание {tour_number}/{task_number}: {exc}")

                # Captains tasks: generate blanks only for captains (always in individual folder)
                if tour.get("captains_task") and is_captain:
                    cap_task_numbers = tour.get("captains_task_numbers") or [1]
                    individual_root = f"Личный зачет/{folder}" if has_team_tours else folder
                    cap_folder = f"{individual_root}/tour_{tour_number}/Задания для капитанов"
                    for cap_task_num in cap_task_numbers:
                        cap_qr_payload = f"attempt:{attempt.id}:tour:{tour_number}:captains_task:{cap_task_num}"
                        try:
                            cap_docx = word_generator.generate_answer_blank(
                                qr_payload=cap_qr_payload,
                                tour_number=tour_number,
                                task_number=int(cap_task_num),
                                mode="Задание для капитанов",
                            )
                            zf.writestr(f"{cap_folder}/задание_{cap_task_num}.docx", cap_docx)
                            added_files += 1
                            for extra_i in range(1, 6):
                                cap_extra = word_generator.generate_answer_blank(
                                    qr_payload=cap_qr_payload,
                                    tour_number=tour_number,
                                    task_number=int(cap_task_num),
                                    mode="Задание для капитанов",
                                    tour_task=f"{tour_number}/cap/{cap_task_num}/{extra_i}",
                                )
                                zf.writestr(f"{cap_folder}/дополнительные бланки/extra_{cap_task_num}_{extra_i}.docx", cap_extra)
                                added_files += 1
                        except Exception as exc:
                            gen_errors.append(f"Задание капитанов {tour_number}/{cap_task_num}: {exc}")

        # --- Badge ---
        entry_token_raw = reg.entry_token.raw_token if reg.entry_token else None
        if entry_token_raw:
            institution_name = participant.institution.name if participant.institution else ""
            last_name, first_name, middle_name = _split_full_name(participant.full_name)
            photo_index = await _load_badge_photo_index(db)
            photo_bytes = _find_badge_photo_bytes(
                photo_index=photo_index,
                city=participant.institution_location,
                institution_name=institution_name,
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name,
            )
            if photo_bytes is None:
                gen_errors.append("Фото для бейджа не найдено — бейдж будет без фотографии")
            try:
                if badge_template is not None:
                    # Use JSON template from visual editor (same as async batch generation)
                    from ....infrastructure.pdf.json_badge_generator import JsonBadgeGenerator
                    json_gen = JsonBadgeGenerator()
                    participant_data = {
                        "LAST_NAME": last_name,
                        "FIRST_NAME": first_name,
                        "MIDDLE_NAME": middle_name,
                        "ROLE": "УЧАСТНИК",
                        "QR_PAYLOAD": entry_token_raw,
                        "PHOTO_BYTES": photo_bytes,
                        "COMPETITION_NAME": competition.name,
                        "INSTITUTION_NAME": institution_name,
                    }
                    badge_pdf = json_gen.generate_badge_pdf(
                        badge_template.config_json or {},
                        participant_data,
                        badge_template.background_image_bytes,
                        on_a4=True,
                    )
                    zf.writestr(f"{folder}/badge.pdf", badge_pdf)
                else:
                    # Fallback: DOCX template from disk
                    badge_docx = word_generator.generate_badge_docx(
                        qr_payload=entry_token_raw,
                        first_name=first_name,
                        last_name=last_name,
                        middle_name=middle_name,
                        role="УЧАСТНИК",
                        participant_school=participant.school,
                        institution_name=institution_name,
                        competition_name=competition.name,
                        photo_bytes=photo_bytes,
                    )
                    zf.writestr(f"{folder}/badge.docx", badge_docx)
                added_files += 1
            except Exception as exc:
                gen_errors.append(f"Бейдж: {exc}")
        else:
            gen_errors.append("Бейдж: отсутствует raw entry token (токен уже использован или не создан)")

    if added_files == 0:
        details = "; ".join(gen_errors) if gen_errors else "Проверьте шаблоны олимпиады."
        raise HTTPException(status_code=400, detail=f"Не удалось сгенерировать файлы: {details}")

    zip_buffer.seek(0)
    encoded_name = quote(f"{participant.full_name}.zip", safe="")
    headers: dict[str, str] = {"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}"}
    if gen_errors:
        headers["X-Warnings"] = quote("; ".join(gen_errors), safe="")
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers=headers,
    )


def _compute_tour_time_item(row: TourTimeModel) -> TourTimeItem:
    """Build TourTimeItem from model row, computing duration_minutes."""
    duration = None
    if row.started_at and row.finished_at and row.finished_at > row.started_at:
        duration = int((row.finished_at - row.started_at).total_seconds() / 60)
    return TourTimeItem(
        tour_number=row.tour_number,
        started_at=row.started_at,
        finished_at=row.finished_at,
        duration_minutes=duration,
    )


@router.get("/competitions/{competition_id}/tour-times", response_model=list[TourTimeItem])
async def get_tour_times(
    competition_id: UUID,
    current_user: User = Depends(require_role(UserRole.ADMIN, UserRole.SCANNER)),
    db: AsyncSession = Depends(get_db),
):
    """Return all recorded tour start/finish times for a competition."""
    result = await db.execute(
        select(TourTimeModel)
        .where(TourTimeModel.competition_id == competition_id)
        .order_by(TourTimeModel.tour_number)
    )
    return [_compute_tour_time_item(row) for row in result.scalars().all()]


@router.put(
    "/competitions/{competition_id}/tour-times/{tour_number}",
    response_model=TourTimeItem,
)
async def set_tour_time(
    competition_id: UUID,
    tour_number: int,
    body: SetTourTimeRequest,
    current_user: User = Depends(require_role(UserRole.ADMIN, UserRole.SCANNER)),
    db: AsyncSession = Depends(get_db),
):
    """Upsert start/finish time for a specific tour of a competition."""
    result = await db.execute(
        select(TourTimeModel).where(
            TourTimeModel.competition_id == competition_id,
            TourTimeModel.tour_number == tour_number,
        )
    )
    row = result.scalar_one_or_none()
    if row is None:
        row = TourTimeModel(
            competition_id=competition_id,
            tour_number=tour_number,
            started_at=body.started_at,
            finished_at=body.finished_at,
        )
        db.add(row)
    else:
        row.started_at = body.started_at
        row.finished_at = body.finished_at
        row.updated_at = datetime.utcnow()
    await db.flush()
    return _compute_tour_time_item(row)


def _build_tour_configs(competition) -> list[TourConfigItem]:
    """Extract tour configs (mode + task_numbers) from competition special_settings."""
    if not competition or not competition.is_special:
        return []
    settings = competition.special_settings or {}
    tours_raw = settings.get("tours")
    if tours_raw and isinstance(tours_raw, list):
        configs = []
        for t in tours_raw:
            if not isinstance(t, dict):
                continue
            configs.append(TourConfigItem(
                tour_number=int(t.get("tour_number", 0)),
                mode=str(t.get("mode", "individual")),
                task_numbers=[int(n) for n in t.get("task_numbers", [])],
                captains_task=bool(t.get("captains_task", False)),
                captains_task_numbers=[int(n) for n in t.get("captains_task_numbers", [])],
            ))
        return configs
    # Fallback: use special_tour_modes without task numbers
    modes = competition.special_tour_modes or []
    return [
        TourConfigItem(tour_number=i + 1, mode=mode, task_numbers=[])
        for i, mode in enumerate(modes)
    ]


@router.get("/competitions/{competition_id}/scoring-progress", response_model=ScoringProgressResponse)
async def get_scoring_progress(
    competition_id: UUID,
    current_user: User = Depends(require_role(UserRole.ADMIN, UserRole.SCANNER)),
    db: AsyncSession = Depends(get_db),
):
    """Return scoring progress for all participants of a competition."""
    use_case = GetScoringProgressUseCase(
        competition_repository=CompetitionRepositoryImpl(db),
        registration_repository=RegistrationRepositoryImpl(db),
        attempt_repository=AttemptRepositoryImpl(db),
        participant_repository=ParticipantRepositoryImpl(db),
    )
    try:
        result = await use_case.execute(competition_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    items = [
        ScoringProgressItem(
            registration_id=item.registration_id,
            participant_id=item.participant_id,
            participant_name=item.participant_name,
            participant_school=item.participant_school,
            variant_number=item.variant_number,
            attempt_id=item.attempt_id,
            attempt_status=item.attempt_status,
            tours=[
                TourProgress(
                    tour_number=t.tour_number,
                    task_scores=t.task_scores,
                    tour_total=t.tour_total,
                    tour_time=t.tour_time,
                )
                for t in item.tours
            ],
            score_total=item.score_total,
            is_captain=item.is_captain,
            captains_task_by_tour=item.captains_task_by_tour,
        )
        for item in result.items
    ]

    # Fetch tour times and include in response
    tt_result = await db.execute(
        select(TourTimeModel)
        .where(TourTimeModel.competition_id == competition_id)
        .order_by(TourTimeModel.tour_number)
    )
    tour_times = [_compute_tour_time_item(row) for row in tt_result.scalars().all()]

    # Build tour configs from competition special_settings
    competition_obj = await CompetitionRepositoryImpl(db).get_by_id(competition_id)
    tour_configs = _build_tour_configs(competition_obj)

    return ScoringProgressResponse(
        competition_id=result.competition_id,
        competition_name=result.competition_name,
        is_special=result.is_special,
        tours_count=result.tours_count,
        items=items,
        total=result.total,
        tour_times=tour_times,
        tour_configs=tour_configs,
    )


@router.get("/competitions/{competition_id}/scoring-progress/export")
async def export_scoring_progress_excel(
    competition_id: UUID,
    current_user: User = Depends(require_role(UserRole.ADMIN, UserRole.SCANNER)),
    db: AsyncSession = Depends(get_db),
):
    """Export scoring progress as a formatted Excel (.xlsx) file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl не установлен") from exc

    # Reuse the same logic as the scoring-progress endpoint
    use_case = GetScoringProgressUseCase(
        competition_repository=CompetitionRepositoryImpl(db),
        registration_repository=RegistrationRepositoryImpl(db),
        attempt_repository=AttemptRepositoryImpl(db),
        participant_repository=ParticipantRepositoryImpl(db),
    )
    try:
        result = await use_case.execute(competition_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    competition_obj = await CompetitionRepositoryImpl(db).get_by_id(competition_id)
    tour_configs = _build_tour_configs(competition_obj)

    tt_result = await db.execute(
        select(TourTimeModel)
        .where(TourTimeModel.competition_id == competition_id)
        .order_by(TourTimeModel.tour_number)
    )
    tour_time_map: dict[int, TourTimeItem] = {
        row.tour_number: _compute_tour_time_item(row)
        for row in tt_result.scalars().all()
    }

    # Mode label mapping
    _MODE_LABELS = {
        "individual": "Личный зачет",
        "individual_captains": "Капитанское задание",
        "team": "Командный зачет",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Результаты"

    header_font = Font(bold=True)
    title_font = Font(bold=True, size=13)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    tour_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _set_header(cell, value):
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    def _set_tour_header(cell, value):
        cell.value = value
        cell.font = header_font
        cell.fill = tour_fill
        cell.alignment = center
        cell.border = border

    competition_name = result.competition_name

    if result.is_special and tour_configs:
        # ---- Build column layout ----
        # Fixed columns: ВУЗ, ФИО, Капитан, Вариант
        fixed_cols = ["ВУЗ", "ФИО", "Капитан", "Вариант"]
        n_fixed = len(fixed_cols)

        # Per tour: task columns + [captain task columns] + итог тура + время тура
        tour_col_spans: list[tuple[int, int]] = []  # (start_col_1based, count)
        tour_time_col_map: dict[int, int] = {}  # tour_number -> 1-based column index
        # captain task columns: tour_number -> list of (cap_task_num, col_index)
        cap_task_col_map: dict[int, list[tuple[int, int]]] = {}
        cap_task_total_col_map: dict[int, int] = {}  # tour_number -> total cap score column index
        col_headers: list[str] = list(fixed_cols)
        for tc in tour_configs:
            start = len(col_headers) + 1
            for task_num in tc.task_numbers:
                col_headers.append(f"Задание {task_num}")
            if tc.captains_task and tc.captains_task_numbers:
                cap_cols: list[tuple[int, int]] = []
                for cap_num in tc.captains_task_numbers:
                    col_idx = len(col_headers) + 1
                    col_headers.append(f"К. задание {cap_num}")
                    cap_cols.append((cap_num, col_idx))
                cap_task_col_map[tc.tour_number] = cap_cols
                cap_task_total_col_map[tc.tour_number] = len(col_headers) + 1
                col_headers.append("Итог к.з.")
            col_headers.append("Итог тура")
            tour_time_col_map[tc.tour_number] = len(col_headers) + 1
            col_headers.append("Время тура")
            n_cap_cols = (len(tc.captains_task_numbers) + 1) if (tc.captains_task and tc.captains_task_numbers) else 0
            span = len(tc.task_numbers) + 2 + n_cap_cols
            tour_col_spans.append((start, span))
        col_headers.append("ИТОГО")
        total_cols = len(col_headers)

        # Row 1: Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=f"Таблица результатов: {competition_name}")
        title_cell.font = title_font
        title_cell.alignment = left

        # Row 2: blank spacer

        # Row 3: Tour group headers (merged)
        for i, tc in enumerate(tour_configs):
            start_c, span = tour_col_spans[i]
            if span > 1:
                ws.merge_cells(start_row=3, start_column=start_c, end_row=3, end_column=start_c + span - 1)
            label = f"Тур {arabic_to_roman(tc.tour_number)} — {_MODE_LABELS.get(tc.mode, tc.mode)}"
            _set_tour_header(ws.cell(row=3, column=start_c), label)
            # Fill remaining merged cells with tour fill
            for c in range(start_c + 1, start_c + span):
                ws.cell(row=3, column=c).fill = tour_fill
                ws.cell(row=3, column=c).border = border
        # Fixed header cells in row 3 (blank with fill)
        for c in range(1, n_fixed + 1):
            ws.cell(row=3, column=c).fill = header_fill
            ws.cell(row=3, column=c).border = border
        # ИТОГО cell in row 3
        ws.cell(row=3, column=total_cols).fill = header_fill
        ws.cell(row=3, column=total_cols).border = border

        # Row 4: Column headers
        for c, name in enumerate(col_headers, start=1):
            _set_header(ws.cell(row=4, column=c), name)

        # Row 5+: Data
        def _parse_tour_time_str(s: str | None):
            """Convert hh.mm.ss string to timedelta, or None."""
            if not s:
                return None
            parts = s.split(".")
            if len(parts) != 3:
                return None
            try:
                from datetime import timedelta as _td
                h, m, sec = int(parts[0]), int(parts[1]), int(parts[2])
                return _td(hours=h, minutes=m, seconds=sec)
            except (ValueError, TypeError):
                return None

        for item in result.items:
            row_data: list = [
                item.participant_school,
                item.participant_name,
                "Да" if item.is_captain else "Нет",
                item.variant_number,
            ]
            for tc in tour_configs:
                tour = next((t for t in item.tours if t.tour_number == tc.tour_number), None)
                for task_num in tc.task_numbers:
                    if tour and tour.task_scores:
                        row_data.append(tour.task_scores.get(str(task_num)))
                    else:
                        row_data.append(None)
                if tc.captains_task and tc.captains_task_numbers:
                    cap_scores = item.captains_task_scores_by_tour.get(tc.tour_number, {})
                    for cap_num in tc.captains_task_numbers:
                        row_data.append(cap_scores.get(str(cap_num)))
                    row_data.append(item.captains_task_by_tour.get(tc.tour_number))
                row_data.append(tour.tour_total if tour else None)
                row_data.append(None)  # time placeholder — written below with number_format
            row_data.append(item.score_total)
            ws.append(row_data)

            # Write time cells with [h]:mm:ss number format
            data_row = ws.max_row
            for tc in tour_configs:
                tour = next((t for t in item.tours if t.tour_number == tc.tour_number), None)
                time_val = _parse_tour_time_str(tour.tour_time if tour else None)
                time_col = tour_time_col_map[tc.tour_number]
                cell = ws.cell(row=data_row, column=time_col)
                cell.value = time_val
                cell.alignment = center
                cell.border = border
                if time_val is not None:
                    cell.number_format = '[h]:mm:ss'

        # Set column widths
        ws.column_dimensions[get_column_letter(1)].width = 30  # ВУЗ
        ws.column_dimensions[get_column_letter(2)].width = 25  # ФИО
        ws.column_dimensions[get_column_letter(3)].width = 9   # Капитан
        ws.column_dimensions[get_column_letter(4)].width = 7   # Вариант
        for c in range(5, total_cols + 1):
            ws.column_dimensions[get_column_letter(c)].width = 11

    else:
        # Non-special: simple layout
        col_headers = ["ВУЗ", "ФИО", "Вариант", "ИТОГО"]
        total_cols = len(col_headers)

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        title_cell = ws.cell(row=1, column=1, value=f"Таблица результатов: {competition_name}")
        title_cell.font = title_font
        title_cell.alignment = left

        for c, name in enumerate(col_headers, start=1):
            _set_header(ws.cell(row=3, column=c), name)

        for item in result.items:
            ws.append([
                item.participant_school,
                item.participant_name,
                item.variant_number,
                item.score_total,
            ])

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 9
        ws.column_dimensions["D"].width = 10

    # Freeze panes below headers
    freeze_row = 5 if (result.is_special and tour_configs) else 4
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = re.sub(r"[^\w\-]", "_", competition_name)[:40]
    filename = f"results_{safe_name}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )


@router.post("/competitions/{competition_id}/results-table/import")
async def import_results_table(
    competition_id: UUID,
    file: UploadFile = File(...),
    current_user: User = Depends(require_role(UserRole.ADMIN, UserRole.SCANNER)),
    db: AsyncSession = Depends(get_db),
):
    """Import results from an XLSX file (same format as exported results table).

    Matches participants by name (column D, "ФИО") and updates their task scores
    per tour. Only non-empty score cells are applied; existing scores are preserved
    if the imported cell is empty.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl не установлен") from exc

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Файл пустой")

    from io import BytesIO
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Не удалось открыть XLSX: {exc}")

    ws = wb.worksheets[0] if wb.worksheets else None
    if not ws:
        raise HTTPException(status_code=400, detail="В файле нет листов")

    # Build tour configs to know column layout
    competition_obj = await CompetitionRepositoryImpl(db).get_by_id(competition_id)
    if not competition_obj:
        raise HTTPException(status_code=404, detail="Олимпиада не найдена")

    tour_configs = _build_tour_configs(competition_obj)
    individual_tours = [tc for tc in tour_configs if tc.mode in ("individual", "individual_captains")]

    # Auto-detect format: scoring-progress export has "ВУЗ" in row 4, column 1
    header_a1 = str(ws.cell(row=4, column=1).value or "").strip()
    is_scoring_progress_format = (header_a1 == "ВУЗ")

    tour_col_map: dict[int, dict] = {}

    if is_scoring_progress_format:
        # Scoring-progress format: A=ВУЗ, B=ФИО, C=Капитан, D=Вариант, then per-tour blocks
        DATA_ROW = 5
        name_col = 2  # column B = ФИО
        ci = 5  # first tour data column (after A, B, C, D)
        for tc in tour_configs:
            n_tasks = len(tc.task_numbers) if tc.task_numbers else 1
            task_cols = list(range(ci, ci + n_tasks))
            total_col_skip = ci + n_tasks
            time_col = ci + n_tasks + 1
            ci += n_tasks + 2  # tasks + total + time
            tour_col_map[tc.tour_number] = {
                "task_cols": task_cols,
                "n_tasks": n_tasks,
                "task_numbers": tc.task_numbers,
                "time_col": time_col,
            }
    else:
        # Results-table format: B=№, C=ВУЗ, D=ФИО, then per-tour blocks
        DATA_ROW = 5
        ci = 2
        ci += 1  # num_col (B)
        inst_col = ci; ci += 1  # C: institution
        name_col = ci; ci += 1  # D: ФИО
        for tc in individual_tours:
            n_tasks = len(tc.task_numbers) if tc.task_numbers else 1
            task_cols = list(range(ci, ci + n_tasks)); ci += n_tasks
            total_col = ci; ci += 1
            time_col = ci; ci += 1
            rank_col = ci; ci += 1
            tour_col_map[tc.tour_number] = {
                "task_cols": task_cols,
                "n_tasks": n_tasks,
                "task_numbers": tc.task_numbers,
                "time_col": time_col,
            }

    # Load all participants with attempts
    use_case = GetScoringProgressUseCase(
        competition_repository=CompetitionRepositoryImpl(db),
        registration_repository=RegistrationRepositoryImpl(db),
        attempt_repository=AttemptRepositoryImpl(db),
        participant_repository=ParticipantRepositoryImpl(db),
    )
    try:
        result = await use_case.execute(competition_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    # Build name → item lookup (normalized)
    def _norm(s: str) -> str:
        return " ".join(s.strip().lower().split())

    name_to_item: dict[str, object] = {}
    for item in result.items:
        name_to_item[_norm(item.participant_name)] = item

    attempt_repo = AttemptRepositoryImpl(db)
    updated_count = 0
    skipped_names: list[str] = []

    for row_idx in range(DATA_ROW, ws.max_row + 1):
        raw_name = ws.cell(row=row_idx, column=name_col).value
        if not raw_name or not isinstance(raw_name, str):
            continue

        # Strip grade suffix like " (9 кл.)"
        clean_name = re.sub(r"\s*\(\d+\s*кл\.\)\s*$", "", raw_name.strip())
        item = name_to_item.get(_norm(clean_name))
        if not item:
            skipped_names.append(clean_name)
            continue

        if not item.attempt_id:
            skipped_names.append(clean_name)
            continue

        attempt = await attempt_repo.get_by_id(item.attempt_id)
        if not attempt:
            continue

        changed = False
        tours_to_import = tour_configs if is_scoring_progress_format else individual_tours
        for tc in tours_to_import:
            tcm = tour_col_map.get(tc.tour_number)
            if not tcm:
                continue

            task_nums = tcm["task_numbers"] or list(range(1, tcm["n_tasks"] + 1))
            scores: dict[int, int] = {}

            for i, col in enumerate(tcm["task_cols"]):
                cell_val = ws.cell(row=row_idx, column=col).value
                if cell_val is not None and isinstance(cell_val, (int, float)):
                    scores[task_nums[i]] = int(cell_val)

            # Parse time from time column
            tour_time_str = None
            time_val = ws.cell(row=row_idx, column=tcm["time_col"]).value
            if time_val is not None:
                from datetime import timedelta
                if isinstance(time_val, timedelta):
                    total_secs = int(time_val.total_seconds())
                    h = total_secs // 3600
                    m = (total_secs % 3600) // 60
                    s = total_secs % 60
                    tour_time_str = f"{h:02d}.{m:02d}.{s:02d}"
                elif isinstance(time_val, str) and re.match(r"^\d{2}\.\d{2}\.\d{2}$", time_val):
                    tour_time_str = time_val
                elif isinstance(time_val, float) and 0.0 <= time_val < 1.0:
                    total_secs = int(round(time_val * 86400))
                    h = total_secs // 3600
                    m = (total_secs % 3600) // 60
                    s = total_secs % 60
                    tour_time_str = f"{h:02d}.{m:02d}.{s:02d}"

            if scores:
                attempt.apply_task_scores(
                    tour_number=tc.tour_number,
                    scores=scores,
                    tour_time=tour_time_str,
                )
                changed = True

        if changed:
            await attempt_repo.update(attempt)
            updated_count += 1

    return {
        "status": "ok",
        "updated": updated_count,
        "skipped": skipped_names,
        "total_rows": ws.max_row - DATA_ROW + 1 if ws.max_row >= DATA_ROW else 0,
    }


@router.get("/competitions/{competition_id}/results-table/export")
async def export_results_table(
    competition_id: UUID,
    current_user: User = Depends(require_role(UserRole.ADMIN, UserRole.SCANNER)),
    db: AsyncSession = Depends(get_db),
):
    """Export final results table as .xlsx with two sheets: personal and team standings."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="openpyxl не установлен") from exc

    use_case = GetScoringProgressUseCase(
        competition_repository=CompetitionRepositoryImpl(db),
        registration_repository=RegistrationRepositoryImpl(db),
        attempt_repository=AttemptRepositoryImpl(db),
        participant_repository=ParticipantRepositoryImpl(db),
    )
    try:
        result = await use_case.execute(competition_id)
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))

    competition_obj = await CompetitionRepositoryImpl(db).get_by_id(competition_id)
    tour_configs = _build_tour_configs(competition_obj)

    tt_result = await db.execute(
        select(TourTimeModel)
        .where(TourTimeModel.competition_id == competition_id)
        .order_by(TourTimeModel.tour_number)
    )
    tour_time_map = {
        row.tour_number: _compute_tour_time_item(row)
        for row in tt_result.scalars().all()
    }

    # Fetch participant grades (not included in use case output)
    from ....infrastructure.database.models import ParticipantModel
    grade_rows = await db.execute(select(ParticipantModel.id, ParticipantModel.grade))
    grade_map: dict[UUID, int | None] = {row.id: row.grade for row in grade_rows}

    competition_name = result.competition_name

    # ── Styles ──────────────────────────────────────────────────────────────
    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    tour_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    calc_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _h(cell, value, fill=None):
        """Write a header cell."""
        cell.value = value
        cell.font = bold
        cell.fill = fill or header_fill
        cell.alignment = center
        cell.border = border

    def _c(cell, value):
        """Write a calculated (formula) cell with grey fill."""
        cell.value = value
        cell.fill = calc_fill
        cell.alignment = center
        cell.border = border

    def _d(cell, value):
        """Write a plain data cell."""
        cell.value = value
        cell.alignment = center
        cell.border = border

    def _dt(cell, value):
        """Write a time data cell (timedelta) with [h]:mm:ss format."""
        cell.value = value
        cell.alignment = center
        cell.border = border
        if value is not None:
            cell.number_format = '[h]:mm:ss'

    def _parse_hms(s: str | None):
        """Convert 'hh.mm.ss' string to timedelta, or None."""
        if not s:
            return None
        parts = s.split(".")
        if len(parts) != 3:
            return None
        try:
            from datetime import timedelta as _td
            h, m, sec = int(parts[0]), int(parts[1]), int(parts[2])
            return _td(hours=h, minutes=m, seconds=sec)
        except (ValueError, TypeError):
            return None

    # ── Helper: tour rank COUNTIFS formula ─────────────────────────────────
    def _tour_rank_formula(total_cl: str, time_cl: str, data_row: int, row: int) -> str:
        return (
            f"=1+COUNTIFS(${total_cl}${data_row}:${total_cl}$9999,"
            f'">"&{total_cl}{row})'
            f"+COUNTIFS(${total_cl}${data_row}:${total_cl}$9999,"
            f'"="&{total_cl}{row},${time_cl}${data_row}:${time_cl}$9999,'
            f'"<"&{time_cl}{row})'
        )

    # ── Helper: final rank COUNTIFS (3-level: rank_sum asc, time asc, score desc) ──
    def _final_rank_formula(rs_cl: str, st_cl: str, ss_cl: str, data_row: int, row: int) -> str:
        return (
            f"=1+COUNTIFS(${rs_cl}${data_row}:${rs_cl}$9999,"
            f'"<"&{rs_cl}{row})'
            f"+COUNTIFS(${rs_cl}${data_row}:${rs_cl}$9999,"
            f'"="&{rs_cl}{row},${st_cl}${data_row}:${st_cl}$9999,'
            f'"<"&{st_cl}{row})'
            f"+COUNTIFS(${rs_cl}${data_row}:${rs_cl}$9999,"
            f'"="&{rs_cl}{row},${st_cl}${data_row}:${st_cl}$9999,'
            f'"="&{st_cl}{row},${ss_cl}${data_row}:${ss_cl}$9999,'
            f'">"&{ss_cl}{row})'
        )

    # ── Helper: team rank (2-level: rank_sum asc, time asc) ────────────────
    def _team_final_rank_formula(rs_cl: str, st_cl: str, data_row: int, row: int) -> str:
        return (
            f"=1+COUNTIFS(${rs_cl}${data_row}:${rs_cl}$9999,"
            f'"<"&{rs_cl}{row})'
            f"+COUNTIFS(${rs_cl}${data_row}:${rs_cl}$9999,"
            f'"="&{rs_cl}{row},${st_cl}${data_row}:${st_cl}$9999,'
            f'"<"&{st_cl}{row})'
        )

    # ── Helper: apply border/fill to a whole row range ────────────────────
    def _fill_row_range(ws, row: int, col_start: int, col_end: int, fill, border):
        for c in range(col_start, col_end + 1):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).border = border

    # ────────────────────────────────────────────────────────────────────────
    # Classify tours
    individual_tours = [tc for tc in tour_configs if tc.mode in ("individual", "individual_captains")]
    captain_tours = [tc for tc in tour_configs if tc.mode == "individual_captains"]
    team_tours = [tc for tc in tour_configs if tc.mode == "team"]

    # Build captain bonus lookup: institution_name → tour_number → total_score
    # Sources: (1) individual_captains tour totals, (2) cap_N task scores from captains_task_by_tour
    captain_bonus: dict[str, dict[int, int | None]] = {}
    for item in result.items:
        if item.is_captain:
            inst = item.participant_school or ""
            for t in item.tours:
                if any(tc.tour_number == t.tour_number for tc in captain_tours):
                    captain_bonus.setdefault(inst, {})[t.tour_number] = t.tour_total
            # Also include cap_N scores from captains_task_by_tour
            for tour_n, cap_score in (item.captains_task_by_tour or {}).items():
                existing = captain_bonus.setdefault(inst, {}).get(tour_n)
                if existing is None:
                    captain_bonus[inst][tour_n] = cap_score
                else:
                    captain_bonus[inst][tour_n] = existing + cap_score

    # Build captain time lookup for team tours: institution_name → tour_number → hh.mm.ss
    captain_time: dict[str, dict[int, str | None]] = {}
    for item in result.items:
        if item.is_captain:
            inst = item.participant_school or ""
            for t in item.tours:
                if any(tc.tour_number == t.tour_number for tc in team_tours) and t.tour_time:
                    captain_time.setdefault(inst, {})[t.tour_number] = t.tour_time

    # Build team tour data: institution_name → tour_number → total_score
    team_tour_data: dict[str, dict[int, int | None]] = {}
    for item in result.items:
        inst = item.participant_school or ""
        for t in item.tours:
            if any(tc.tour_number == t.tour_number for tc in team_tours):
                prev = team_tour_data.setdefault(inst, {}).get(t.tour_number) or 0
                addition = t.tour_total or 0
                team_tour_data.setdefault(inst, {})[t.tour_number] = prev + addition

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1: Личный зачет
    # ════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Личный зачет"

    DATA_ROW_S1 = 5  # data starts at row 5

    # ── Build column index map ──────────────────────────────────────────────
    # col indices are 1-based; column A is skipped (reference uses B as first)
    ci = 2  # start at column B

    num_col = ci; ci += 1           # B: №
    inst_col = ci; ci += 1          # C: institution (SUMIF key)
    vid_col = ci; ci += 1           # D: ФИО / Вид

    # Per individual tour: task_cols, total_col, time_col, rank_col
    tour_col_map: dict[int, dict] = {}
    for tc in individual_tours:
        n_tasks = len(tc.task_numbers) if tc.task_numbers else 1
        task_cols = list(range(ci, ci + n_tasks)); ci += n_tasks
        total_col = ci; ci += 1
        time_col = ci; ci += 1
        rank_col = ci; ci += 1
        tour_col_map[tc.tour_number] = {
            "task_cols": task_cols,
            "total_col": total_col,
            "time_col": time_col,
            "rank_col": rank_col,
            "n_tasks": n_tasks,
            "task_numbers": tc.task_numbers,
        }

    # Summary columns
    rank_sum_col = ci; ci += 1
    time_sum_col = ci; ci += 1
    score_sum_col = ci; ci += 1
    final_rank_col = ci

    total_cols_s1 = final_rank_col

    # ── Rows 1–2: title ───────────────────────────────────────────────────
    ws1.merge_cells(start_row=1, start_column=2, end_row=1, end_column=total_cols_s1)
    t = ws1.cell(row=1, column=2, value="Личный зачет")
    t.font = title_font; t.alignment = left
    ws1.merge_cells(start_row=2, start_column=2, end_row=2, end_column=total_cols_s1)
    t2 = ws1.cell(row=2, column=2, value=competition_name)
    t2.font = Font(bold=True, size=11); t2.alignment = left

    # ── Row 3: group headers ──────────────────────────────────────────────
    _fill_row_range(ws1, 3, 2, total_cols_s1, header_fill, border)
    ws1.merge_cells(start_row=3, start_column=2, end_row=4, end_column=num_col)
    _h(ws1.cell(row=3, column=num_col), "№ п/п")
    ws1.merge_cells(start_row=3, start_column=inst_col, end_row=3, end_column=vid_col)
    _h(ws1.cell(row=3, column=inst_col), "Участники олимпиады")

    for tc in individual_tours:
        tcm = tour_col_map[tc.tour_number]
        span_start = tcm["task_cols"][0]
        span_end = tcm["rank_col"]
        if span_end > span_start:
            ws1.merge_cells(start_row=3, start_column=span_start, end_row=3, end_column=span_end)
        _h(ws1.cell(row=3, column=span_start), f"Тур {arabic_to_roman(tc.tour_number)}", fill=tour_fill)
        _fill_row_range(ws1, 3, span_start + 1, span_end, tour_fill, border)

    if individual_tours:
        ws1.merge_cells(start_row=3, start_column=rank_sum_col, end_row=3, end_column=score_sum_col)
        _h(ws1.cell(row=3, column=rank_sum_col), "Итого")

    ws1.merge_cells(start_row=3, start_column=final_rank_col, end_row=4, end_column=final_rank_col)
    _h(ws1.cell(row=3, column=final_rank_col), "Место с окончательным порядком")

    # ── Row 4: column headers ─────────────────────────────────────────────
    _fill_row_range(ws1, 4, 2, total_cols_s1, header_fill, border)
    _h(ws1.cell(row=4, column=inst_col), "Наименование команды (ВУЗ)")
    _h(ws1.cell(row=4, column=vid_col), "ФИО")

    for tc in individual_tours:
        tcm = tour_col_map[tc.tour_number]
        task_nums = tcm["task_numbers"] or list(range(1, tcm["n_tasks"] + 1))
        for i, col in enumerate(tcm["task_cols"]):
            _h(ws1.cell(row=4, column=col), f"Задание {task_nums[i]}")
        _h(ws1.cell(row=4, column=tcm["total_col"]), "Итого баллов")
        _h(ws1.cell(row=4, column=tcm["time_col"]), "Время выполнения")
        _h(ws1.cell(row=4, column=tcm["rank_col"]), "Место")

    if individual_tours:
        _h(ws1.cell(row=4, column=rank_sum_col), "Сумма мест")
        _h(ws1.cell(row=4, column=time_sum_col), "Суммарное время")
        _h(ws1.cell(row=4, column=score_sum_col), "Сумма баллов")

    # ── Rows 5+: data ────────────────────────────────────────────────────
    for idx, item in enumerate(result.items):
        row = DATA_ROW_S1 + idx
        grade = grade_map.get(item.participant_id)
        name_val = item.participant_name or ""
        if grade:
            name_val = f"{name_val} ({grade} кл.)"

        # Row number
        if idx == 0:
            _c(ws1.cell(row=row, column=num_col), 1)
        else:
            _c(ws1.cell(row=row, column=num_col), f"={get_column_letter(num_col)}{row - 1}+1")

        # Institution (SUMIF key)
        _d(ws1.cell(row=row, column=inst_col), item.participant_school or "")
        # ФИО + класс
        _d(ws1.cell(row=row, column=vid_col), name_val)

        # Tour data
        for tc in individual_tours:
            tcm = tour_col_map[tc.tour_number]
            tour_prog = next((t for t in item.tours if t.tour_number == tc.tour_number), None)
            task_nums = tcm["task_numbers"] or list(range(1, tcm["n_tasks"] + 1))

            # Task scores (input cells)
            for i, col in enumerate(tcm["task_cols"]):
                score = None
                if tour_prog and tour_prog.task_scores:
                    score = tour_prog.task_scores.get(str(task_nums[i]))
                _d(ws1.cell(row=row, column=col), score)

            # Tour total (formula)
            task_letters = "+".join(
                f"{get_column_letter(c)}{row}" for c in tcm["task_cols"]
            )
            _c(ws1.cell(row=row, column=tcm["total_col"]), f"={task_letters}")

            # Tour time (per-participant if available, else competition-wide)
            time_val = None
            if tour_prog and tour_prog.tour_time:
                # Parse hh.mm.ss format to timedelta
                from datetime import timedelta
                parts = tour_prog.tour_time.split(".")
                if len(parts) == 3:
                    try:
                        h, m, s = int(parts[0]), int(parts[1]), int(parts[2])
                        time_val = timedelta(hours=h, minutes=m, seconds=s)
                    except ValueError:
                        pass
            if time_val is None:
                tt = tour_time_map.get(tc.tour_number)
                if tt and tt.duration_minutes is not None:
                    from datetime import timedelta
                    time_val = timedelta(minutes=tt.duration_minutes)
            _dt(ws1.cell(row=row, column=tcm["time_col"]), time_val)

            # Tour rank (formula)
            total_cl = get_column_letter(tcm["total_col"])
            time_cl = get_column_letter(tcm["time_col"])
            _c(ws1.cell(row=row, column=tcm["rank_col"]),
               _tour_rank_formula(total_cl, time_cl, DATA_ROW_S1, row))

        # Summary formulas
        if individual_tours:
            rank_cols_str = "+".join(
                f"{get_column_letter(tour_col_map[tc.tour_number]['rank_col'])}{row}"
                for tc in individual_tours
            )
            _c(ws1.cell(row=row, column=rank_sum_col), f"={rank_cols_str}")

            time_cols_str = "+".join(
                f"{get_column_letter(tour_col_map[tc.tour_number]['time_col'])}{row}"
                for tc in individual_tours
            )
            _c(ws1.cell(row=row, column=time_sum_col), f"={time_cols_str}")
            ws1.cell(row=row, column=time_sum_col).number_format = '[h]:mm:ss'

            score_cols_str = "+".join(
                f"{get_column_letter(tour_col_map[tc.tour_number]['total_col'])}{row}"
                for tc in individual_tours
            )
            _c(ws1.cell(row=row, column=score_sum_col), f"={score_cols_str}")

            rs_cl = get_column_letter(rank_sum_col)
            st_cl = get_column_letter(time_sum_col)
            ss_cl = get_column_letter(score_sum_col)
            _c(ws1.cell(row=row, column=final_rank_col),
               _final_rank_formula(rs_cl, st_cl, ss_cl, DATA_ROW_S1, row))
        else:
            # No individual tours: use score_total directly
            _d(ws1.cell(row=row, column=final_rank_col), item.score_total)

    # ── Column widths & freeze ────────────────────────────────────────────
    ws1.column_dimensions[get_column_letter(num_col)].width = 5
    ws1.column_dimensions[get_column_letter(inst_col)].width = 30
    ws1.column_dimensions[get_column_letter(vid_col)].width = 25
    for tc in individual_tours:
        tcm = tour_col_map[tc.tour_number]
        for col in tcm["task_cols"]:
            ws1.column_dimensions[get_column_letter(col)].width = 11
        ws1.column_dimensions[get_column_letter(tcm["total_col"])].width = 13
        ws1.column_dimensions[get_column_letter(tcm["time_col"])].width = 13
        ws1.column_dimensions[get_column_letter(tcm["rank_col"])].width = 8
    if individual_tours:
        ws1.column_dimensions[get_column_letter(rank_sum_col)].width = 11
        ws1.column_dimensions[get_column_letter(time_sum_col)].width = 13
        ws1.column_dimensions[get_column_letter(score_sum_col)].width = 13
        ws1.column_dimensions[get_column_letter(final_rank_col)].width = 10

    ws1.freeze_panes = ws1.cell(row=DATA_ROW_S1, column=2)
    ws1.auto_filter.ref = f"{get_column_letter(inst_col)}4:{get_column_letter(final_rank_col)}4"

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2: Командный зачет  (only if competition has tours)
    # ════════════════════════════════════════════════════════════════════════
    if tour_configs:
        ws2 = wb.create_sheet("Командный зачет")
        DATA_ROW_S2 = 7  # data starts at row 7 (reference: rows 1-6 are headers)

        # Collect unique institutions in order of first appearance
        seen: dict[str, None] = {}
        for item in result.items:
            inst = item.participant_school or ""
            seen[inst] = None
        institutions = list(seen.keys())

        # ── Column index map for sheet 2 ──────────────────────────────────
        ci2 = 2
        num2_col = ci2; ci2 += 1       # B: №
        inst2_col = ci2; ci2 += 1      # C: institution

        tour2_col_map: dict[int, dict] = {}

        # Individual tour columns in sheet 2: bonus(if captains), total(SUMIF), time(SUMIF), rank
        for tc in individual_tours:
            bonus_col = None
            if tc.mode == "individual_captains":
                bonus_col = ci2; ci2 += 1
            total2_col = ci2; ci2 += 1
            time2_col = ci2; ci2 += 1
            rank2_col = ci2; ci2 += 1
            tour2_col_map[tc.tour_number] = {
                "total_col": total2_col,
                "time_col": time2_col,
                "rank_col": rank2_col,
                "bonus_col": bonus_col,
                "mode": tc.mode,
            }

        # Team tour columns: total, time, rank
        team2_col_map: dict[int, dict] = {}
        for tc in team_tours:
            total2_col = ci2; ci2 += 1
            time2_col = ci2; ci2 += 1
            rank2_col = ci2; ci2 += 1
            team2_col_map[tc.tour_number] = {
                "total_col": total2_col,
                "time_col": time2_col,
                "rank_col": rank2_col,
            }

        rank2_sum_col = ci2; ci2 += 1
        time2_sum_col = ci2; ci2 += 1
        final2_rank_col = ci2
        total_cols_s2 = final2_rank_col

        # ── Rows 1–2: title ───────────────────────────────────────────────
        ws2.merge_cells(start_row=1, start_column=2, end_row=1, end_column=total_cols_s2)
        t = ws2.cell(row=1, column=2, value="Командный зачет")
        t.font = title_font; t.alignment = left
        ws2.merge_cells(start_row=2, start_column=2, end_row=2, end_column=total_cols_s2)
        t2 = ws2.cell(row=2, column=2, value=competition_name)
        t2.font = Font(bold=True, size=11); t2.alignment = left

        # ── Rows 3–6: headers ─────────────────────────────────────────────
        _fill_row_range(ws2, 3, 2, total_cols_s2, header_fill, border)
        _fill_row_range(ws2, 4, 2, total_cols_s2, header_fill, border)
        _fill_row_range(ws2, 5, 2, total_cols_s2, header_fill, border)
        _fill_row_range(ws2, 6, 2, total_cols_s2, header_fill, border)

        # Rows 3-6 span for №, institution
        ws2.merge_cells(start_row=3, start_column=num2_col, end_row=6, end_column=num2_col)
        _h(ws2.cell(row=3, column=num2_col), "№ п/п")
        ws2.merge_cells(start_row=3, start_column=inst2_col, end_row=6, end_column=inst2_col)
        _h(ws2.cell(row=3, column=inst2_col), "Наименование команды (ВУЗ)")

        # Group headers (row 5) and column headers (row 6) for individual tours
        for tc in individual_tours:
            tm = tour2_col_map[tc.tour_number]
            span_start = tm["bonus_col"] if tm["bonus_col"] else tm["total_col"]
            span_end = tm["rank_col"]
            ws2.merge_cells(start_row=5, start_column=span_start, end_row=5, end_column=span_end)
            _h(ws2.cell(row=5, column=span_start), f"Тур {arabic_to_roman(tc.tour_number)}", fill=tour_fill)
            _fill_row_range(ws2, 5, span_start + 1, span_end, tour_fill, border)

            if tm["bonus_col"]:
                _h(ws2.cell(row=6, column=tm["bonus_col"]), "Доп. задание (капитан)")
            _h(ws2.cell(row=6, column=tm["total_col"]), "Итого баллов")
            _h(ws2.cell(row=6, column=tm["time_col"]), "Время выполнения")
            _h(ws2.cell(row=6, column=tm["rank_col"]), "Место")

        for tc in team_tours:
            tm = team2_col_map[tc.tour_number]
            ws2.merge_cells(
                start_row=5, start_column=tm["total_col"],
                end_row=5, end_column=tm["rank_col"]
            )
            _h(ws2.cell(row=5, column=tm["total_col"]), f"Тур {arabic_to_roman(tc.tour_number)} (командный)", fill=tour_fill)
            _fill_row_range(ws2, 5, tm["total_col"] + 1, tm["rank_col"], tour_fill, border)
            _h(ws2.cell(row=6, column=tm["total_col"]), "Итого баллов")
            _h(ws2.cell(row=6, column=tm["time_col"]), "Время выполнения")
            _h(ws2.cell(row=6, column=tm["rank_col"]), "Место")

        # Summary header
        ws2.merge_cells(
            start_row=5, start_column=rank2_sum_col,
            end_row=5, end_column=total_cols_s2
        )
        _h(ws2.cell(row=5, column=rank2_sum_col), "Итого")
        _h(ws2.cell(row=6, column=rank2_sum_col), "Сумма мест")
        _h(ws2.cell(row=6, column=time2_sum_col), "Суммарное время")
        _h(ws2.cell(row=6, column=final2_rank_col), "Итоговое место")

        # ── Rows 7+: data ─────────────────────────────────────────────────
        s1_name = "Личный зачет"
        inst_cl_s1 = get_column_letter(inst_col)

        for idx, inst in enumerate(institutions):
            row = DATA_ROW_S2 + idx

            if idx == 0:
                _c(ws2.cell(row=row, column=num2_col), 1)
            else:
                _c(ws2.cell(row=row, column=num2_col),
                   f"={get_column_letter(num2_col)}{row - 1}+1")

            _d(ws2.cell(row=row, column=inst2_col), inst)

            all_rank_cols = []
            all_time_cols = []

            # Individual tours: SUMIF formulas
            for tc in individual_tours:
                tm = tour2_col_map[tc.tour_number]
                s1_total_cl = get_column_letter(tour_col_map[tc.tour_number]["total_col"])
                s1_time_cl = get_column_letter(tour_col_map[tc.tour_number]["time_col"])
                team_inst_ref = f"C{row}"

                sumif_total = (
                    f"=SUMIF('{s1_name}'!${inst_cl_s1}:${inst_cl_s1},"
                    f"{team_inst_ref},"
                    f"'{s1_name}'!${s1_total_cl}:${s1_total_cl})"
                )
                sumif_time = (
                    f"=SUMIF('{s1_name}'!${inst_cl_s1}:${inst_cl_s1},"
                    f"{team_inst_ref},"
                    f"'{s1_name}'!${s1_time_cl}:${s1_time_cl})"
                )

                # If captain bonus: total includes bonus column
                if tm["bonus_col"]:
                    bonus_val = captain_bonus.get(inst, {}).get(tc.tour_number)
                    _d(ws2.cell(row=row, column=tm["bonus_col"]), bonus_val)
                    bonus_ref = f"{get_column_letter(tm['bonus_col'])}{row}"
                    _c(ws2.cell(row=row, column=tm["total_col"]),
                       f"{sumif_total}+{bonus_ref}")
                else:
                    _c(ws2.cell(row=row, column=tm["total_col"]), sumif_total)

                _c(ws2.cell(row=row, column=tm["time_col"]), sumif_time)
                ws2.cell(row=row, column=tm["time_col"]).number_format = '[h]:mm:ss'

                # Rank for this tour
                total2_cl = get_column_letter(tm["total_col"])
                time2_cl = get_column_letter(tm["time_col"])
                _c(ws2.cell(row=row, column=tm["rank_col"]),
                   _tour_rank_formula(total2_cl, time2_cl, DATA_ROW_S2, row))

                all_rank_cols.append(get_column_letter(tm["rank_col"]))
                all_time_cols.append(get_column_letter(tm["time_col"]))

            # Team tours: direct values + captain task bonus
            for tc in team_tours:
                tm = team2_col_map[tc.tour_number]
                team_total_val = team_tour_data.get(inst, {}).get(tc.tour_number)
                cap_bonus = captain_bonus.get(inst, {}).get(tc.tour_number) or 0
                total_with_bonus = (team_total_val or 0) + cap_bonus
                _d(ws2.cell(row=row, column=tm["total_col"]),
                   total_with_bonus if (team_total_val is not None or cap_bonus) else None)

                # Prefer captain's per-participant tour time; fall back to competition-wide map
                ct_str = captain_time.get(inst, {}).get(tc.tour_number)
                if ct_str:
                    time_val = _parse_hms(ct_str)
                else:
                    tt = tour_time_map.get(tc.tour_number)
                    time_val = None
                    if tt and tt.duration_minutes is not None:
                        from datetime import timedelta
                        time_val = timedelta(minutes=tt.duration_minutes)
                _dt(ws2.cell(row=row, column=tm["time_col"]), time_val)

                total2_cl = get_column_letter(tm["total_col"])
                time2_cl = get_column_letter(tm["time_col"])
                _c(ws2.cell(row=row, column=tm["rank_col"]),
                   _tour_rank_formula(total2_cl, time2_cl, DATA_ROW_S2, row))

                all_rank_cols.append(get_column_letter(tm["rank_col"]))
                all_time_cols.append(get_column_letter(tm["time_col"]))

            # Summary
            if all_rank_cols:
                rank_sum_str = "+".join(f"{cl}{row}" for cl in all_rank_cols)
                _c(ws2.cell(row=row, column=rank2_sum_col), f"={rank_sum_str}")

                time_sum_str = "+".join(f"{cl}{row}" for cl in all_time_cols)
                _c(ws2.cell(row=row, column=time2_sum_col), f"={time_sum_str}")
                ws2.cell(row=row, column=time2_sum_col).number_format = '[h]:mm:ss'

                rs2_cl = get_column_letter(rank2_sum_col)
                st2_cl = get_column_letter(time2_sum_col)
                _c(ws2.cell(row=row, column=final2_rank_col),
                   _team_final_rank_formula(rs2_cl, st2_cl, DATA_ROW_S2, row))

        # ── Column widths & freeze ─────────────────────────────────────────
        ws2.column_dimensions[get_column_letter(num2_col)].width = 5
        ws2.column_dimensions[get_column_letter(inst2_col)].width = 30
        for tm in tour2_col_map.values():
            ws2.column_dimensions[get_column_letter(tm["total_col"])].width = 13
            ws2.column_dimensions[get_column_letter(tm["time_col"])].width = 13
            ws2.column_dimensions[get_column_letter(tm["rank_col"])].width = 8
            if tm["bonus_col"]:
                ws2.column_dimensions[get_column_letter(tm["bonus_col"])].width = 18
        for tm in team2_col_map.values():
            ws2.column_dimensions[get_column_letter(tm["total_col"])].width = 13
            ws2.column_dimensions[get_column_letter(tm["time_col"])].width = 13
            ws2.column_dimensions[get_column_letter(tm["rank_col"])].width = 8
        ws2.column_dimensions[get_column_letter(rank2_sum_col)].width = 11
        ws2.column_dimensions[get_column_letter(time2_sum_col)].width = 13
        ws2.column_dimensions[get_column_letter(final2_rank_col)].width = 12

        ws2.freeze_panes = ws2.cell(row=DATA_ROW_S2, column=2)

    # ── Stream response ───────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = re.sub(r"[^\w\-]", "_", competition_name)[:40]
    filename = f"results_table_{safe_name}.xlsx"
    encoded_filename = quote(filename)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )


@router.post("/registrations/{competition_id}/badges-pdf/start")
async def start_badges_pdf(
    competition_id: UUID,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Start asynchronous badge PDF generation. Returns a task_id to poll for status."""
    from ....infrastructure.database.models import CompetitionModel
    from ....infrastructure.tasks.badge_tasks import generate_badges_pdf

    comp_result = await db.execute(
        select(CompetitionModel).where(CompetitionModel.id == competition_id)
    )
    if not comp_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Олимпиада не найдена")

    task = generate_badges_pdf.delay(str(competition_id))
    return {"task_id": task.id}


@router.get("/badge-tasks/{task_id}/status")
async def get_badge_task_status(
    task_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    """Poll the status of a badge PDF generation task."""
    from ....infrastructure.tasks.celery_app import celery_app as _celery

    result = _celery.AsyncResult(task_id)
    state = result.state  # PENDING, STARTED, PROGRESS, SUCCESS, FAILURE
    info = result.info or {}

    if state == "SUCCESS":
        payload = info if isinstance(info, dict) else {}
        return {
            "state": "SUCCESS",
            "status": payload.get("status", "completed"),
            "count": payload.get("count"),
            "object_name": payload.get("object_name"),
        }
    if state == "FAILURE":
        return {"state": "FAILURE", "message": str(info)}
    if state == "PROGRESS" and isinstance(info, dict):
        return {
            "state": "PROGRESS",
            "stage": info.get("stage", ""),
            "current": info.get("current", 0),
            "total": info.get("total", 0),
        }
    return {"state": state}


@router.get("/badge-tasks/{task_id}/download")
async def download_badge_task_pdf(
    task_id: str,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    """Download the generated badge PDF once the task is complete."""
    from ....infrastructure.tasks.celery_app import celery_app as _celery
    from ....infrastructure.storage import MinIOStorage
    from ....infrastructure.tasks.badge_tasks import BADGE_PDF_BUCKET

    result = _celery.AsyncResult(task_id)
    if result.state != "SUCCESS":
        raise HTTPException(status_code=400, detail=f"Задача ещё не завершена (состояние: {result.state})")

    payload = result.result or {}
    if not isinstance(payload, dict) or payload.get("status") != "completed":
        message = payload.get("message", "неизвестная ошибка") if isinstance(payload, dict) else str(payload)
        raise HTTPException(status_code=500, detail=f"Генерация завершилась с ошибкой: {message}")

    object_name = payload.get("object_name")
    if not object_name:
        raise HTTPException(status_code=500, detail="Результат задачи не содержит пути к файлу")

    try:
        storage = MinIOStorage()
        pdf_bytes = storage.download_file(bucket=BADGE_PDF_BUCKET, object_name=object_name)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Не удалось загрузить PDF из хранилища: {exc}") from exc

    filename = object_name.rsplit("/", 1)[-1]
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/registrations/{competition_id}/badges-pdf")
async def download_badges_pdf(
    competition_id: UUID,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Download template-based badge PDF grouped by institution."""
    from ....infrastructure.database.models import (
        RegistrationModel,
        CompetitionModel,
        ParticipantModel,
    )
    from ....infrastructure.pdf.badge_template_pdf_generator import (
        BadgeTemplatePdfGenerator,
        TemplateBadgePdfItem,
    )
    from io import BytesIO

    comp_result = await db.execute(
        select(CompetitionModel).where(CompetitionModel.id == competition_id)
    )
    competition = comp_result.scalar_one_or_none()
    if not competition:
        raise HTTPException(status_code=404, detail="Олимпиада не найдена")

    stmt = (
        select(RegistrationModel)
        .where(RegistrationModel.competition_id == competition_id)
        .options(
            selectinload(RegistrationModel.entry_token),
            selectinload(RegistrationModel.participant).selectinload(ParticipantModel.institution),
        )
        .order_by(RegistrationModel.created_at)
    )
    result = await db.execute(stmt)
    registrations = result.scalars().all()

    word_generator = WordTemplateGenerator()
    if not word_generator.is_docx_to_pdf_available():
        raise HTTPException(
            status_code=500,
            detail="Для генерации PDF бейджей из DOCX-шаблона требуется LibreOffice (soffice).",
        )
    photo_index = await _load_badge_photo_index(db)

    prepared: list[dict[str, Any]] = []
    for reg in registrations:
        participant = reg.participant
        if not participant:
            continue

        entry_token_raw = None
        if reg.entry_token and reg.entry_token.raw_token:
            entry_token_raw = reg.entry_token.raw_token
        if not entry_token_raw:
            continue

        institution_name = ""
        if participant.institution:
            institution_name = participant.institution.name or ""

        last_name, first_name, middle_name = _split_full_name(participant.full_name)
        photo_bytes = _find_badge_photo_bytes(
            photo_index=photo_index,
            city=participant.institution_location,
            institution_name=institution_name,
            last_name=last_name,
            first_name=first_name,
            middle_name=middle_name,
        )

        prepared.append(
            {
                "institution": institution_name,
                "full_name": participant.full_name,
                "docx": word_generator.generate_badge_docx(
                    qr_payload=entry_token_raw,
                    first_name=first_name,
                    last_name=last_name,
                    middle_name=middle_name,
                    role="УЧАСТНИК",
                    participant_school=participant.school,
                    institution_name=institution_name,
                    competition_name=competition.name,
                    photo_bytes=photo_bytes,
                ),
            }
        )

    prepared.sort(key=lambda item: (item["institution"] or "", item["full_name"] or ""))
    if not prepared:
        raise HTTPException(status_code=400, detail="Нет зарегистрированных участников для генерации бейджей")

    docx_files: dict[str, bytes] = {}
    name_to_institution: dict[str, str] = {}
    for index, item in enumerate(prepared, start=1):
        file_name = f"badge_{index:04d}.docx"
        docx_files[file_name] = item["docx"]
        name_to_institution[file_name] = item["institution"]

    try:
        converted_pdf = word_generator.convert_docx_files_to_pdf(docx_files)
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=f"Не удалось конвертировать бейджи в PDF: {exc}") from exc
    template_items: list[TemplateBadgePdfItem] = []
    for file_name in docx_files:
        template_items.append(
            TemplateBadgePdfItem(
                institution=name_to_institution[file_name],
                pdf_bytes=converted_pdf[file_name],
            )
        )

    generator = BadgeTemplatePdfGenerator()
    pdf_bytes = generator.generate_grouped_pdf(competition.name, template_items)

    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="badges_{competition_id}.pdf"'
        },
    )

@router.get("/registrations/{competition_id}/badges-docx")
async def download_badges_docx(
    competition_id: UUID,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Download badge DOCX files generated from editable Word template."""
    from ....infrastructure.database.models import (
        RegistrationModel,
        CompetitionModel,
        ParticipantModel,
    )
    from io import BytesIO

    comp_result = await db.execute(
        select(CompetitionModel).where(CompetitionModel.id == competition_id)
    )
    competition = comp_result.scalar_one_or_none()
    if not competition:
        raise HTTPException(status_code=404, detail="Олимпиада не найдена")

    stmt = (
        select(RegistrationModel)
        .where(RegistrationModel.competition_id == competition_id)
        .options(
            selectinload(RegistrationModel.entry_token),
            selectinload(RegistrationModel.participant).selectinload(ParticipantModel.institution),
        )
        .order_by(RegistrationModel.created_at)
    )
    result = await db.execute(stmt)
    registrations = result.scalars().all()

    word_generator = WordTemplateGenerator()
    template_paths = word_generator.get_template_paths()
    photo_index = await _load_badge_photo_index(db)
    zip_buffer = io.BytesIO()
    index = 1

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        try:
            zf.write(template_paths["badge"], arcname="_templates/badge_template.docx")
        except Exception:  # noqa: BLE001
            pass

        for reg in registrations:
            participant = reg.participant
            if not participant:
                continue

            entry_token_raw = None
            if reg.entry_token and reg.entry_token.raw_token:
                entry_token_raw = reg.entry_token.raw_token
            if not entry_token_raw:
                continue

            institution_name = ""
            if participant.institution:
                institution_name = participant.institution.name or ""

            last_name, first_name, middle_name = _split_full_name(participant.full_name)
            photo_bytes = _find_badge_photo_bytes(
                photo_index=photo_index,
                city=participant.institution_location,
                institution_name=institution_name,
                last_name=last_name,
                first_name=first_name,
                middle_name=middle_name,
            )
            badge_docx = word_generator.generate_badge_docx(
                qr_payload=entry_token_raw,
                first_name=first_name,
                last_name=last_name,
                middle_name=middle_name,
                role="УЧАСТНИК",
                participant_school=participant.school,
                institution_name=institution_name,
                competition_name=competition.name,
                photo_bytes=photo_bytes,
            )

            institution_slug = _slugify_folder_name(institution_name or "without_institution")
            participant_slug = _slugify_folder_name(participant.full_name)
            if not participant_slug:
                participant_slug = str(participant.id)
            filename = f"{institution_slug}/{index:03d}_{participant_slug}.docx"
            zf.writestr(filename, badge_docx)
            index += 1

    zip_buffer.seek(0)
    return StreamingResponse(
        BytesIO(zip_buffer.getvalue()),
        media_type="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="badges_{competition_id}_docx.zip"'
        },
    )


@router.get("/competitions/{competition_id}/seating-plan")
async def get_seating_plan(
    competition_id: UUID,
    tour_number: int | None = Query(None, ge=1, description="Tour number for special seating mode"),
    current_user: User = Depends(require_role(UserRole.ADMIN, UserRole.ADMITTER, UserRole.INVIGILATOR)),
    db: AsyncSession = Depends(get_db),
):
    """Get seating plan grouped by room for viewing/printing."""
    from ....infrastructure.database.models import (
        CompetitionModel,
        RoomModel,
        SeatAssignmentModel,
        RegistrationModel,
        ParticipantModel,
        InstitutionModel,
    )

    competition_result = await db.execute(
        select(CompetitionModel).where(CompetitionModel.id == competition_id)
    )
    competition = competition_result.scalar_one_or_none()
    if not competition:
        raise HTTPException(status_code=404, detail="Олимпиада не найдена")

    await _ensure_seat_assignments_for_competition(competition=competition, db=db)

    default_seat_matrix_columns = _resolve_default_seat_matrix_columns(competition)
    special_tour_context = _resolve_special_tour_context(competition, tour_number=tour_number)
    is_team_mode = bool(special_tour_context and special_tour_context["is_team_mode"])
    selected_tour_number = int(special_tour_context["tour_number"]) if special_tour_context else None

    rooms_result = await db.execute(
        select(RoomModel)
        .where(RoomModel.competition_id == competition_id)
        .order_by(RoomModel.name.asc())
    )
    rooms = rooms_result.scalars().all()

    plan_rooms: list[dict[str, Any]] = []
    for room in rooms:
        seats_result = await db.execute(
            select(
                SeatAssignmentModel.seat_number,
                SeatAssignmentModel.variant_number,
                ParticipantModel.full_name,
                ParticipantModel.institution_id,
                ParticipantModel.institution_location,
                ParticipantModel.is_captain,
                InstitutionModel.name.label("institution_name"),
            )
            .join(RegistrationModel, RegistrationModel.id == SeatAssignmentModel.registration_id)
            .join(ParticipantModel, ParticipantModel.id == RegistrationModel.participant_id)
            .outerjoin(InstitutionModel, InstitutionModel.id == ParticipantModel.institution_id)
            .where(SeatAssignmentModel.room_id == room.id)
            .order_by(SeatAssignmentModel.seat_number.asc())
        )
        seat_rows = seats_result.all()
        seats = [
            {
                "seat_number": row.seat_number,
                "variant_number": row.variant_number,
                "participant_name": row.full_name,
                "institution_id": row.institution_id,
                "institution_name": row.institution_name,
                "institution_location": row.institution_location,
                "is_captain": row.is_captain,
            }
            for row in seat_rows
        ]
        seats_by_number = {seat["seat_number"]: seat for seat in seats}
        seats_per_table = _resolve_room_seats_per_table(
            competition=competition,
            room_id=room.id,
            is_team_mode=is_team_mode,
        )
        room_seat_matrix_columns = _resolve_room_seat_matrix_columns(
            competition=competition,
            room_id=room.id,
            is_team_mode=is_team_mode,
        )
        room_tables = _build_room_tables(
            room_capacity=room.capacity,
            seats_by_number=seats_by_number,
            seats_per_table=seats_per_table,
        )
        merge_groups = _resolve_room_team_table_merges(
            competition=competition,
            room_id=room.id,
            tour_number=selected_tour_number if is_team_mode else None,
        )
        normalized_merge_groups = _annotate_tables_with_merges(room_tables, merge_groups)
        if is_team_mode and normalized_merge_groups:
            _project_team_seating_for_merges(room_tables, normalized_merge_groups)
            projected_seats = [
                seat
                for table in room_tables
                for seat in table.get("seats", [])
                if seat.get("occupied")
            ]
            seats = sorted(projected_seats, key=lambda seat: int(seat.get("seat_number") or 0))
            seats_by_number = {seat["seat_number"]: seat for seat in seats}

        matrix_rows_count = (room.capacity + room_seat_matrix_columns - 1) // room_seat_matrix_columns
        seat_matrix: list[list[dict[str, Any]]] = []
        for matrix_row in range(matrix_rows_count):
            row_cells: list[dict[str, Any]] = []
            for matrix_col in range(room_seat_matrix_columns):
                seat_number = matrix_row * room_seat_matrix_columns + matrix_col + 1
                if seat_number > room.capacity:
                    continue
                seat_data = seats_by_number.get(seat_number)
                row_cells.append(
                    {
                        "seat_number": seat_number,
                        "table_number": ((seat_number - 1) // seats_per_table) + 1,
                        "seat_at_table": ((seat_number - 1) % seats_per_table) + 1,
                        "occupied": seat_data is not None,
                        "variant_number": seat_data["variant_number"] if seat_data else None,
                        "participant_name": seat_data["participant_name"] if seat_data else None,
                        "institution_id": seat_data["institution_id"] if seat_data else None,
                        "institution_name": seat_data["institution_name"] if seat_data else None,
                        "institution_location": seat_data["institution_location"] if seat_data else None,
                        "is_captain": seat_data["is_captain"] if seat_data else False,
                    }
                )
            if row_cells:
                seat_matrix.append(row_cells)

        plan_rooms.append(
            {
                "room_id": str(room.id),
                "room_name": room.name,
                "capacity": room.capacity,
                "occupied": len(seats),
                "seat_matrix_columns": room_seat_matrix_columns,
                "seats_per_table": seats_per_table,
                "tables_count": len(room_tables),
                "occupied_tables": sum(1 for table in room_tables if table["occupied"]),
                "table_merges": normalized_merge_groups,
                "has_table_merges": bool(normalized_merge_groups),
                "tables": room_tables,
                "seat_matrix": seat_matrix,
                "seats": seats,
            }
        )

    return {
        "competition_id": str(competition.id),
        "competition_name": competition.name,
        "tour_number": special_tour_context["tour_number"] if special_tour_context else None,
        "tour_mode": special_tour_context["mode"] if special_tour_context else None,
        "is_team_mode": is_team_mode,
        "seat_matrix_columns": default_seat_matrix_columns,
        "rooms": plan_rooms,
    }


@router.get("/competitions/{competition_id}/seating-plan/print", response_class=HTMLResponse)
async def print_seating_plan(
    competition_id: UUID,
    tour_number: int | None = Query(None, ge=1, description="Tour number for special seating mode"),
    current_user: User = Depends(require_role(UserRole.ADMIN, UserRole.ADMITTER, UserRole.INVIGILATOR)),
    db: AsyncSession = Depends(get_db),
):
    """Render printable seating plan as HTML."""
    seating = await get_seating_plan(
        competition_id=competition_id,
        tour_number=tour_number,
        current_user=current_user,
        db=db,
    )

    mode_label_map = {
        "individual": "Индивидуальный",
        "individual_captains": "Индивидуальный (капитаны)",
        "team": "Командный",
    }
    mode_label = mode_label_map.get(seating.get("tour_mode"), seating.get("tour_mode") or "-")

    def esc(value: Any, fallback: str = "-") -> str:
        if value is None:
            return fallback
        text = str(value).strip()
        if not text:
            return fallback
        return escape(text)

    room_sections: list[str] = []
    for room in seating["rooms"]:
        tables = room.get("tables") or []
        table_by_number: dict[int, dict[str, Any]] = {
            int(table.get("table_number", 0)): table
            for table in tables
            if int(table.get("table_number", 0)) > 0
        }
        table_numbers = sorted(table_by_number.keys())
        tables_count = len(table_numbers)

        # Horizontal columns: if config says 3, we render 3 desk columns.
        desk_cols = max(int(room.get("seat_matrix_columns", seating.get("seat_matrix_columns", 3))), 1)
        desk_rows = max((tables_count + desk_cols - 1) // desk_cols, 1)

        def _table_position(table_number: int) -> tuple[int, int]:
            index = table_number - 1
            row = (index // desk_cols) + 1
            col = (index % desk_cols) + 1
            return row, col

        def _render_half(seats: list[dict[str, Any]]) -> str:
            if not seats:
                return "<div class='seat-card empty'>Свободно</div>"

            seat_cards: list[str] = []
            for seat in seats:
                if not seat.get("occupied"):
                    seat_cards.append("<div class='seat-card empty'>Свободно</div>")
                    continue
                captain_badge = "<div class='seat-flag'>Капитан</div>" if seat.get("is_captain") else ""
                seat_cards.append(
                    "<div class='seat-card'>"
                    f"<div class='seat-name'>{esc(seat.get('participant_name'))}</div>"
                    f"<div class='seat-meta'>Место {esc(seat.get('seat_number'))} | Вариант {esc(seat.get('variant_number'))}</div>"
                    f"<div class='seat-meta'>{esc(seat.get('institution_location'))}</div>"
                    f"{captain_badge}"
                    "</div>"
                )
            return "".join(seat_cards)

        def _render_table_inner(table: dict[str, Any], title_prefix: str = "Стол") -> str:
            seats = sorted(
                table.get("seats", []),
                key=lambda s: int(s.get("seat_at_table") or 0),
            )
            if not seats:
                left_half: list[dict[str, Any]] = []
                right_half: list[dict[str, Any]] = []
            else:
                split_index = max((len(seats) + 1) // 2, 1)
                left_half = seats[:split_index]
                right_half = seats[split_index:]
            return (
                f"<div class='desk-head'>{title_prefix} {table.get('table_number')}</div>"
                "<div class='desk-split'>"
                f"<div class='desk-half'>{_render_half(left_half)}</div>"
                f"<div class='desk-half'>{_render_half(right_half)}</div>"
                "</div>"
            )

        normalized_groups: list[list[int]] = []
        for raw_group in (room.get("table_merges") or []):
            if not isinstance(raw_group, list):
                continue
            group = sorted({int(num) for num in raw_group if int(num) in table_by_number})
            if len(group) >= 2:
                normalized_groups.append(group)

        rendered_tables: set[int] = set()
        desk_blocks: list[str] = []
        invalid_merges: list[str] = []

        for group in normalized_groups:
            positions = [(num, *_table_position(num)) for num in group]
            rows = sorted({row for _, row, _ in positions})
            cols = sorted({col for _, _, col in positions})

            is_horizontal_chain = (
                len(rows) == 1
                and cols == list(range(min(cols), max(cols) + 1))
                and len(cols) == len(group)
            )
            is_vertical_chain = (
                len(cols) == 1
                and rows == list(range(min(rows), max(rows) + 1))
                and len(rows) == len(group)
            )

            if not (is_horizontal_chain or is_vertical_chain):
                invalid_merges.append("+".join(str(num) for num in group))
                continue

            if is_horizontal_chain:
                row_start = rows[0]
                col_start = min(cols)
                col_span = len(cols)
                grid_style = f"grid-column:{col_start} / span {col_span};grid-row:{row_start};"
                stack_class = "merged-stack horizontal"
            else:
                col_start = cols[0]
                row_start = min(rows)
                row_span = len(rows)
                grid_style = f"grid-column:{col_start};grid-row:{row_start} / span {row_span};"
                stack_class = "merged-stack vertical"

            grouped_tables = [table_by_number[num] for num in group]
            seats_total = sum(len(table.get("seats", [])) for table in grouped_tables)
            occupied_total = sum(
                1
                for table in grouped_tables
                for seat in table.get("seats", [])
                if seat.get("occupied")
            )
            merged_label = "+".join(str(num) for num in group)
            merged_state = "occupied" if occupied_total else "free"
            merged_cards = "".join(
                f"<div class='merged-table-card'>{_render_table_inner(table_by_number[num])}</div>"
                for num in group
            )
            desk_blocks.append(
                f"<div class='desk merged {merged_state}' style='{grid_style}'>"
                f"<div class='desk-merge-title'>Столы {merged_label} ({occupied_total}/{seats_total})</div>"
                f"<div class='{stack_class}'>{merged_cards}</div>"
                "</div>"
            )
            rendered_tables.update(group)

        for table_number in table_numbers:
            if table_number in rendered_tables:
                continue
            table = table_by_number[table_number]
            row, col = _table_position(table_number)
            state = "occupied" if any(seat.get("occupied") for seat in table.get("seats", [])) else "free"
            desk_blocks.append(
                f"<div class='desk {state}' style='grid-column:{col};grid-row:{row};'>"
                f"{_render_table_inner(table)}"
                "</div>"
            )

        auditorium_html = (
            "<div class='auditorium'>"
            f"<div class='room-center'>{esc(room['room_name'], fallback='')}</div>"
            f"<div class='auditorium-grid' style='--aud-rows:{desk_rows};--aud-cols:{desk_cols};'>"
            f"{''.join(desk_blocks) or '<div class=\"no-desks\">Нет столов</div>'}"
            "</div>"
            "</div>"
        )

        rows_html: list[str] = []
        for seat in room["seats"]:
            rows_html.append(
                "<tr>"
                f"<td>{seat['seat_number']}</td>"
                f"<td>{esc(seat['participant_name'])}</td>"
                f"<td>{esc(seat.get('institution_name'))}</td>"
                f"<td>{esc(seat.get('institution_location'))}</td>"
                f"<td>{'Да' if seat.get('is_captain') else 'Нет'}</td>"
                f"<td>{esc(seat.get('variant_number'))}</td>"
                "</tr>"
            )

        table_html = (
            "<table>"
            "<thead><tr>"
            "<th>Место</th><th>Участник</th><th>Учреждение</th><th>Город/филиал</th><th>Капитан</th><th>Вариант</th>"
            "</tr></thead>"
            f"<tbody>{''.join(rows_html) or '<tr><td colspan=\"6\">Нет рассадки</td></tr>'}</tbody>"
            "</table>"
        )

        merges_text = ", ".join("+".join(str(table_number) for table_number in group) for group in normalized_groups)
        merges_line = f"<div class='muted'>Объединения столов: {escape(merges_text)}</div>" if merges_text else ""
        invalid_merges_line = (
            f"<div class='merge-warning'>Не удалось отрисовать объединение для групп: {escape(', '.join(invalid_merges))}</div>"
            if invalid_merges
            else ""
        )

        room_sections.append(
            "<section>"
            f"<h2>{esc(room['room_name'], fallback='Аудитория')} ({room['occupied']}/{room['capacity']})</h2>"
            f"<div class='muted'>Колонок столов: {desk_cols} | Рядов столов: {desk_rows}</div>"
            f"<div class='muted'>Мест за столом: {room.get('seats_per_table', 1)} | Столов: {room.get('tables_count', 0)} (занято {room.get('occupied_tables', 0)})</div>"
            f"{merges_line}"
            f"{invalid_merges_line}"
            f"{auditorium_html}"
            f"{table_html}"
            "</section>"
        )

    html = (
        "<!doctype html>"
        "<html><head><meta charset='utf-8'>"
        f"<title>Схема рассадки - {esc(seating['competition_name'])}</title>"
        "<style>"
        "body{font-family:Arial,sans-serif;margin:24px;color:#111;}"
        "h1{margin-bottom:8px;} h2{margin-top:26px;margin-bottom:6px;}"
        ".muted{color:#5f6368;font-size:12px;}"
        ".merge-warning{margin-top:4px;color:#b3261e;font-size:12px;}"
        ".auditorium{position:relative;border:2px solid #202124;border-radius:10px;margin-top:12px;padding:16px;background:linear-gradient(180deg,#ffffff 0%,#f8f9fa 100%);min-height:300px;}"
        ".room-center{position:absolute;inset:0;display:flex;align-items:center;justify-content:center;font-size:34px;font-weight:700;color:rgba(95,99,104,.25);pointer-events:none;letter-spacing:1px;text-transform:uppercase;}"
        ".auditorium-grid{position:relative;z-index:1;display:grid;grid-template-columns:repeat(var(--aud-cols),minmax(95px,1fr));grid-template-rows:repeat(var(--aud-rows),minmax(58px,1fr));gap:10px;min-height:260px;}"
        ".desk{border:1.5px solid #5f6368;border-radius:8px;padding:6px;background:#fff;display:flex;flex-direction:column;gap:6px;line-height:1.25;}"
        ".desk.occupied{background:#e8f0fe;border-color:#1a73e8;}"
        ".desk.free{background:#f8f9fa;border-color:#9aa0a6;}"
        ".desk.merged{border-width:2.5px;background:#e6f4ea;border-color:#137333;padding:8px;}"
        ".desk-merge-title{font-size:12px;font-weight:700;text-align:center;}"
        ".merged-stack{display:flex;gap:0;align-items:stretch;}"
        ".merged-stack.vertical{flex-direction:column;}"
        ".merged-stack.horizontal{flex-direction:row;}"
        ".merged-table-card{border:1px solid #7aa786;border-radius:0;padding:4px;background:#fff;flex:1 1 0;}"
        ".merged-stack.horizontal .merged-table-card + .merged-table-card{margin-left:-1px;}"
        ".merged-stack.vertical .merged-table-card + .merged-table-card{margin-top:-1px;}"
        ".merged-stack.horizontal .merged-table-card:first-child{border-top-left-radius:6px;border-bottom-left-radius:6px;}"
        ".merged-stack.horizontal .merged-table-card:last-child{border-top-right-radius:6px;border-bottom-right-radius:6px;}"
        ".merged-stack.vertical .merged-table-card:first-child{border-top-left-radius:6px;border-top-right-radius:6px;}"
        ".merged-stack.vertical .merged-table-card:last-child{border-bottom-left-radius:6px;border-bottom-right-radius:6px;}"
        ".desk-head{font-size:12px;font-weight:700;text-align:center;margin-bottom:4px;}"
        ".desk-split{display:grid;grid-template-columns:1fr 1fr;border:1px solid #9aa0a6;border-radius:6px;overflow:hidden;background:#fff;}"
        ".desk-half{min-height:74px;padding:4px;display:flex;flex-direction:column;justify-content:flex-start;}"
        ".desk-half + .desk-half{border-left:1px solid #9aa0a6;}"
        ".seat-card{font-size:10px;padding:2px 0;border-bottom:1px dashed #d2d6dc;}"
        ".seat-card:last-child{border-bottom:none;}"
        ".seat-card.empty{display:flex;align-items:center;justify-content:center;min-height:18px;color:#6b7280;}"
        ".seat-name{font-weight:700;word-break:break-word;}"
        ".seat-meta{color:#374151;word-break:break-word;}"
        ".seat-flag{display:inline-block;margin-top:2px;font-size:9px;font-weight:700;color:#0b57d0;}"
        ".no-desks{grid-column:1 / -1;display:flex;align-items:center;justify-content:center;color:#5f6368;border:1px dashed #9aa0a6;border-radius:8px;min-height:64px;}"
        "table{border-collapse:collapse;width:100%;margin-top:10px;}"
        "th,td{border:1px solid #c7c7c7;padding:6px 8px;font-size:12px;text-align:left;}"
        "th{background:#f1f3f4;}"
        "@media print{body{margin:0.5cm;} section{break-inside:avoid;page-break-inside:avoid;} .auditorium{break-inside:avoid;}}"
        "</style></head><body>"
        "<h1>Схема рассадки</h1>"
        f"<div><strong>Олимпиада:</strong> {esc(seating['competition_name'])}</div>"
        f"<div class='muted'>Тур: {esc(arabic_to_roman(int(seating['tour_number'])) if seating.get('tour_number') else '—')} | Режим: {esc(mode_label)}</div>"
        f"{''.join(room_sections)}"
        "</body></html>"
    )

    return HTMLResponse(content=html)


# --- Special Olympiad ---


@router.get("/special/templates")
async def get_special_templates_info(
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    """Get Word template paths used for special olympiad documents."""
    generator = WordTemplateGenerator()
    paths = generator.get_template_paths()
    templates = []
    for kind in ("answer_blank", "a3_cover", "badge"):
        path = Path(paths[kind])
        info: dict[str, Any] = {"kind": kind, "path": str(path), "filename": path.name}
        # Read original uploaded filename from sidecar .meta file
        meta_path = Path(f"{paths[kind]}.meta")
        if meta_path.exists():
            try:
                info["display_filename"] = meta_path.read_text(encoding="utf-8").strip()
            except Exception:
                info["display_filename"] = path.name if path.exists() else "Нет шаблона"
        else:
            info["display_filename"] = "Нет шаблона"
        if path.exists():
            stat = path.stat()
            info["size_bytes"] = stat.st_size
            info["modified_at"] = datetime.fromtimestamp(stat.st_mtime).isoformat()
        templates.append(info)
    return {"templates": templates}


@router.get("/special/templates/{template_kind}/download")
async def download_special_template(
    template_kind: str,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    """Download current DOCX template."""
    generator = WordTemplateGenerator()
    paths = generator.get_template_paths()

    if template_kind not in paths:
        raise HTTPException(status_code=404, detail="Неизвестный тип шаблона")

    path = paths[template_kind]
    try:
        with open(path, "rb") as f:
            content = f.read()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Не удалось открыть шаблон: {exc}")

    filename = Path(path).name
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
    )


@router.post("/special/templates/{template_kind}/upload")
async def upload_special_template(
    template_kind: str,
    file: UploadFile = File(...),
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    """Upload replacement DOCX template."""
    generator = WordTemplateGenerator()
    paths = generator.get_template_paths()

    if template_kind not in paths:
        raise HTTPException(status_code=404, detail="Неизвестный тип шаблона")
    if not (file.filename or "").lower().endswith(".docx"):
        raise HTTPException(status_code=400, detail="Нужен файл .docx")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Файл пустой")

    path = paths[template_kind]
    try:
        with open(path, "wb") as f:
            f.write(content)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Не удалось сохранить шаблон: {exc}")

    # Save original filename to sidecar .meta file
    meta_path = Path(f"{path}.meta")
    try:
        meta_path.write_text(file.filename or path, encoding="utf-8")
    except Exception:
        pass

    return {"status": "ok", "template_kind": template_kind, "path": path}


@router.delete("/special/templates/{template_kind}")
async def delete_special_template(
    template_kind: str,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    """Delete a custom DOCX template, restoring the default."""
    generator = WordTemplateGenerator()
    paths = generator.get_template_paths()

    if template_kind not in paths:
        raise HTTPException(status_code=404, detail="Неизвестный тип шаблона")

    path = Path(paths[template_kind])
    if path.exists():
        path.unlink()

    # Remove sidecar .meta file
    meta_path = Path(f"{paths[template_kind]}.meta")
    if meta_path.exists():
        meta_path.unlink()

    # Recreate default template
    generator.ensure_templates_exist()
    return {"status": "ok", "template_kind": template_kind, "message": "Шаблон сброшен на стандартный"}


@router.post("/special/templates/badge/photos/upload")
async def upload_badge_photos_zip(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Upload ZIP archive with participant photos for badge template token {{PHOTO}}."""
    from ....infrastructure.database.models import BadgePhotoModel

    if not (file.filename or "").lower().endswith(".zip"):
        raise HTTPException(status_code=400, detail="Нужен ZIP-архив с фотографиями")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Файл пустой")

    content_type_by_ext = {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
    }
    records: dict[str, tuple[str, bytes, str]] = {}
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for info in zf.infolist():
                if info.is_dir():
                    continue
                ext = Path(info.filename).suffix.lower()
                if ext not in content_type_by_ext:
                    continue

                raw_path = str(info.filename).replace("\\", "/")
                raw_parts = [part for part in raw_path.split("/") if part]
                if not raw_parts or any(part == ".." for part in raw_parts):
                    continue

                normalized_key = WordTemplateGenerator._normalize_photo_key(  # type: ignore[attr-defined]
                    str(Path(raw_path).with_suffix("")).replace("\\", "/")
                )
                if not normalized_key:
                    continue

                image_bytes = zf.read(info)
                if not image_bytes:
                    continue

                records[normalized_key] = (
                    raw_path,
                    image_bytes,
                    content_type_by_ext[ext],
                )
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Некорректный ZIP-архив")
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Не удалось обработать архив фотографий: {exc}")

    await db.execute(delete(BadgePhotoModel))
    for normalized_key, (original_path, image_bytes, content_type) in records.items():
        db.add(
            BadgePhotoModel(
                normalized_key=normalized_key,
                original_path=original_path,
                content_type=content_type,
                image_bytes=image_bytes,
            )
        )
    imported = len(records)

    return {
        "status": "ok",
        "imported_files": imported,
        "imported": imported,
        "path_hint": "Город/Учреждение/Фамилия_Имя_Отчество.png",
    }


@router.post("/special/templates/badge/fonts/upload")
async def upload_badge_fonts(
    file: UploadFile = File(...),
    current_user: User = Depends(require_role(UserRole.ADMIN)),
):
    """Upload TTF/OTF font files (single file or ZIP) used during LibreOffice badge PDF conversion.

    Fonts are stored on disk in the templates/word/fonts/ directory and are automatically
    passed to LibreOffice when generating badge PDFs, so custom fonts like Magistral are
    rendered correctly without requiring system-wide installation in the container.
    """
    word_generator = WordTemplateGenerator()
    word_generator.ensure_templates_exist()
    fonts_dir = word_generator.fonts_dir

    filename = (file.filename or "").lower()
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Файл пустой")

    imported = 0

    if filename.endswith(".zip"):
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as zf:
                for info in zf.infolist():
                    if info.is_dir():
                        continue
                    ext = Path(info.filename).suffix.lower()
                    if ext not in {".ttf", ".otf"}:
                        continue
                    font_name = Path(info.filename).name
                    if not font_name:
                        continue
                    (fonts_dir / font_name).write_bytes(zf.read(info))
                    imported += 1
        except zipfile.BadZipFile:
            raise HTTPException(status_code=400, detail="Некорректный ZIP-архив")
    elif filename.endswith(".ttf") or filename.endswith(".otf"):
        font_name = Path(file.filename or "font").name
        (fonts_dir / font_name).write_bytes(content)
        imported = 1
    else:
        raise HTTPException(
            status_code=400,
            detail="Поддерживаются файлы .ttf, .otf или ZIP-архив с ними",
        )

    # Install fonts system-wide immediately so LibreOffice can use them
    # on the very next badge PDF generation without requiring a restart.
    word_generator.install_fonts_system_wide()

    return {"status": "ok", "imported_files": imported}


@router.post("/competitions/{competition_id}/special/import-participants")
async def import_special_participants(
    competition_id: UUID,
    file: UploadFile = File(...),
    register_to_competition: bool = Query(True, description="Register imported participants to this competition"),
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Import participants for a special olympiad from JSON/CSV/XLSX."""
    competition_repo = CompetitionRepositoryImpl(db)
    competition = await competition_repo.get_by_id(competition_id)
    if not competition:
        raise HTTPException(status_code=404, detail="Олимпиада не найдена")
    if not competition.is_special:
        raise HTTPException(status_code=400, detail="Импорт доступен только для олимпиад с пометкой 'особая'")

    file_name = file.filename or ""
    file_bytes = await file.read()
    if not file_bytes:
        raise HTTPException(status_code=400, detail="Файл пустой")

    try:
        rows = _parse_import_file(file_name, file_bytes)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))

    if not rows:
        return {
            "total_rows": 0,
            "created_users": 0,
            "created_participants": 0,
            "registered_to_competition": 0,
            "skipped": 0,
            "errors": [],
        }

    from ....domain.entities import Institution, Participant

    user_repo = UserRepositoryImpl(db)
    participant_repo = ParticipantRepositoryImpl(db)
    institution_repo = InstitutionRepositoryImpl(db)
    registration_repo = RegistrationRepositoryImpl(db)
    entry_token_repo = EntryTokenRepositoryImpl(db)
    register_uc = RegisterForCompetitionUseCase(
        registration_repository=registration_repo,
        competition_repository=competition_repo,
        participant_repository=participant_repo,
        entry_token_repository=entry_token_repo,
        token_service=TokenService(settings.hmac_secret_key),
    )

    summary = {
        "total_rows": len(rows),
        "created_users": 0,
        "created_participants": 0,
        "registered_to_competition": 0,
        "skipped": 0,
        "errors": [],
    }

    for idx, row in enumerate(rows, start=1):
        try:
            normalized = _normalize_record(row)
            full_name = str(normalized.get("full_name") or "").strip()
            institution_name = str(normalized.get("institution") or "").strip()

            if len(full_name) < 2:
                raise ValueError("Поле ФИО обязательно")
            if len(institution_name) < 2:
                raise ValueError("Поле ВУЗ/учреждение обязательно")

            institution_location_raw = normalized.get("institution_location")
            institution_location = (
                str(institution_location_raw).strip()
                if institution_location_raw is not None and str(institution_location_raw).strip()
                else None
            )
            is_captain = _parse_bool(normalized.get("is_captain"))
            dob = _parse_dob(normalized.get("dob"))

            email = str(normalized.get("email") or "").strip().lower()
            if not email:
                email = derive_imported_email(full_name, dob)

            def _opt_str(key: str) -> str | None:
                value = normalized.get(key)
                if value is None:
                    return None
                text = str(value).strip()
                return text or None

            position = _opt_str("position")
            military_rank = _opt_str("military_rank")
            passport_series_number = _opt_str("passport_series_number")
            passport_issued_by = _opt_str("passport_issued_by")
            passport_issued_date = _parse_dob(normalized.get("passport_issued_date"))
            military_booklet_number = _opt_str("military_booklet_number")
            military_personal_number = _opt_str("military_personal_number")

            institution = await institution_repo.get_by_name(institution_name)
            if not institution:
                institution = await institution_repo.create(
                    Institution(
                        id=uuid4(),
                        name=institution_name,
                        city=institution_location,
                    )
                )

            user = await user_repo.get_by_email(email)
            participant = None

            if user is None:
                generated_password = secrets.token_urlsafe(12)
                user = await user_repo.create(
                    User(
                        id=uuid4(),
                        email=email,
                        password_hash=hash_password(generated_password),
                        role=UserRole.PARTICIPANT,
                        is_active=True,
                    )
                )
                summary["created_users"] += 1
            elif user.role != UserRole.PARTICIPANT:
                raise ValueError(f"Email {email} уже занят пользователем с ролью {user.role.value}")

            team_name = f"{institution_name} ({institution_location})" if institution_location else institution_name
            participant = await participant_repo.get_by_user_id(user.id)
            if participant is None:
                participant = await participant_repo.create(
                    Participant(
                        id=uuid4(),
                        user_id=user.id,
                        full_name=full_name,
                        school=team_name,
                        grade=None,
                        institution_id=institution.id,
                        institution_location=institution_location,
                        is_captain=is_captain,
                        dob=dob,
                        position=position,
                        military_rank=military_rank,
                        passport_series_number=passport_series_number,
                        passport_issued_by=passport_issued_by,
                        passport_issued_date=passport_issued_date,
                        military_booklet_number=military_booklet_number,
                        military_personal_number=military_personal_number,
                    )
                )
                summary["created_participants"] += 1
            else:
                participant.update_profile(
                    full_name=full_name,
                    school=team_name,
                    institution_location=institution_location,
                    is_captain=is_captain,
                    dob=dob,
                    position=position,
                    military_rank=military_rank,
                    passport_series_number=passport_series_number,
                    passport_issued_by=passport_issued_by,
                    passport_issued_date=passport_issued_date,
                    military_booklet_number=military_booklet_number,
                    military_personal_number=military_personal_number,
                )
                participant.institution_id = institution.id
                await participant_repo.update(participant)

            if register_to_competition:
                try:
                    await register_uc.execute(
                        participant_id=participant.id,
                        competition_id=competition_id,
                        skip_status_check=True,
                    )
                    summary["registered_to_competition"] += 1
                except ValueError as exc:
                    if "уже зарегистрированы" in str(exc):
                        summary["skipped"] += 1
                    else:
                        raise

        except Exception as exc:  # noqa: BLE001
            summary["errors"].append({"row": idx, "error": str(exc)})

    return summary


@router.post("/competitions/{competition_id}/special/admit-all-and-download")
async def admit_all_special_and_download(
    competition_id: UUID,
    request: Request,
    current_user: User = Depends(require_role(UserRole.ADMIN)),
    db: AsyncSession = Depends(get_db),
):
    """Admit all pending participants in special olympiad and return ZIP with sheets by participant folders."""
    competition_repo = CompetitionRepositoryImpl(db)
    competition = await competition_repo.get_by_id(competition_id)
    if not competition:
        raise HTTPException(status_code=404, detail="Олимпиада не найдена")
    if not competition.is_special:
        raise HTTPException(status_code=400, detail="Операция доступна только для олимпиад с пометкой 'особая'")

    from ....infrastructure.database.models import (
        RegistrationModel,
        AnswerSheetModel,
    )

    registration_stmt = (
        select(RegistrationModel)
        .where(RegistrationModel.competition_id == competition_id)
        .options(
            selectinload(RegistrationModel.participant),
            selectinload(RegistrationModel.entry_token),
            selectinload(RegistrationModel.attempts),
        )
        .order_by(RegistrationModel.created_at.asc())
    )
    registrations_result = await db.execute(registration_stmt)
    registrations = registrations_result.scalars().all()

    approve_uc = ApproveAdmissionUseCase(
        token_service=TokenService(settings.hmac_secret_key),
        entry_token_repository=EntryTokenRepositoryImpl(db),
        registration_repository=RegistrationRepositoryImpl(db),
        competition_repository=competition_repo,
        attempt_repository=AttemptRepositoryImpl(db),
        audit_log_repository=AuditLogRepositoryImpl(db),
        answer_sheet_repository=AnswerSheetRepositoryImpl(db),
        storage=MinIOStorage(),
        sheet_generator=SheetGenerator(),
        room_repository=RoomRepositoryImpl(db),
        seat_assignment_repository=SeatAssignmentRepositoryImpl(db),
        participant_repository=ParticipantRepositoryImpl(db),
    )

    admitted_now = 0
    admit_errors: list[dict[str, Any]] = []

    # 1) Admit pending registrations.
    for reg in registrations:
        if reg.status.value != "pending":
            continue
        if not reg.entry_token or not reg.entry_token.raw_token:
            admit_errors.append(
                {
                    "registration_id": str(reg.id),
                    "participant": reg.participant.full_name if reg.participant else "—",
                    "error": "У регистрации отсутствует raw entry token",
                }
            )
            continue

        try:
            await approve_uc.execute(
                registration_id=reg.id,
                raw_entry_token=reg.entry_token.raw_token,
                admitter_user_id=current_user.id,
                ip_address=request.client.host if request.client else None,
            )
            admitted_now += 1
        except Exception as exc:  # noqa: BLE001
            admit_errors.append(
                {
                    "registration_id": str(reg.id),
                    "participant": reg.participant.full_name if reg.participant else "—",
                    "error": str(exc),
                }
            )

    # 2) Reload registrations to include created attempts.
    registrations_result = await db.execute(registration_stmt)
    registrations = registrations_result.scalars().all()

    # 3) Build ZIP archive (DOCX templates + legacy PDFs).
    storage = MinIOStorage()
    word_generator = WordTemplateGenerator()
    template_paths = word_generator.get_template_paths()
    tours = _extract_special_tours(competition)
    has_team_tours = any(str(t["mode"]) == "team" for t in tours)
    zip_buffer = io.BytesIO()
    added_files = 0
    mode_labels = {
        "individual": "Индивидуальный",
        "individual_captains": "Индивидуальный (капитаны)",
        "team": "Командный",
    }

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        # Include editable templates in archive for quick customization.
        try:
            zf.write(template_paths["answer_blank"], arcname="_templates/special_answer_blank_template.docx")
            zf.write(template_paths["a3_cover"], arcname="_templates/special_cover_a3_template.docx")
            added_files += 2
        except Exception:  # noqa: BLE001
            pass

        for reg in registrations:
            participant = reg.participant
            if not participant:
                continue
            if not reg.attempts:
                continue

            attempt = reg.attempts[0]
            folder = _slugify_folder_name(f"{participant.full_name}_{participant.id}")
            # When team tours are present, individual sheets go under "Личный зачет/"
            individual_root = f"Личный зачет/{folder}" if has_team_tours else folder

            sheets_stmt = (
                select(AnswerSheetModel)
                .where(AnswerSheetModel.attempt_id == attempt.id)
                .order_by(AnswerSheetModel.created_at.asc())
            )
            sheets_result = await db.execute(sheets_stmt)
            sheets = sheets_result.scalars().all()

            # Generate DOCX set based on editable Word templates (individual/captains tours only).
            for tour in tours:
                tour_number = int(tour["tour_number"])
                mode = str(tour["mode"])
                mode_label = mode_labels.get(mode, mode)
                task_numbers = tour["task_numbers"]

                # Team tours are handled separately after this loop
                if mode == "team":
                    continue

                cover_qr_payload = f"attempt:{attempt.id}:tour:{tour_number}:cover"
                try:
                    cover_docx = word_generator.generate_a3_cover(
                        qr_payload=cover_qr_payload,
                        tour_number=tour_number,
                        mode=mode_label,
                    )
                    zf.writestr(f"{individual_root}/tour_{tour_number}/A3_tour_{tour_number}.docx", cover_docx)
                    added_files += 1
                except Exception as exc:  # noqa: BLE001
                    admit_errors.append(
                        {
                            "registration_id": str(reg.id),
                            "participant": participant.full_name,
                            "error": f"A3 tour {tour_number}: {exc}",
                        }
                    )

                for task_number in task_numbers:
                    task_qr_payload = f"attempt:{attempt.id}:tour:{tour_number}:task:{task_number}"
                    try:
                        task_docx = word_generator.generate_answer_blank(
                            qr_payload=task_qr_payload,
                            tour_number=tour_number,
                            task_number=int(task_number),
                            mode=mode_label,
                        )
                        task_folder = f"{individual_root}/tour_{tour_number}/task_{task_number}"
                        zf.writestr(
                            f"{task_folder}/task_{task_number}.docx",
                            task_docx,
                        )
                        added_files += 1
                        for extra_i in range(1, 6):
                            extra_docx = word_generator.generate_answer_blank(
                                qr_payload=task_qr_payload,
                                tour_number=tour_number,
                                task_number=int(task_number),
                                mode=mode_label,
                                tour_task=f"{tour_number}/{task_number}/{extra_i}",
                            )
                            zf.writestr(
                                f"{task_folder}/дополнительные бланки/extra_{extra_i}.docx",
                                extra_docx,
                            )
                            added_files += 1
                    except Exception as exc:  # noqa: BLE001
                        admit_errors.append(
                            {
                                "registration_id": str(reg.id),
                                "participant": participant.full_name,
                                "error": f"Task {tour_number}/{task_number}: {exc}",
                            }
                        )

                # Captains tasks: generate blanks only for captains (always in individual folder)
                if tour.get("captains_task") and participant.is_captain:
                    cap_task_numbers = tour.get("captains_task_numbers") or [1]
                    cap_folder = f"{individual_root}/tour_{tour_number}/Задания для капитанов"
                    for cap_task_num in cap_task_numbers:
                        cap_qr_payload = f"attempt:{attempt.id}:tour:{tour_number}:captains_task:{cap_task_num}"
                        try:
                            cap_docx = word_generator.generate_answer_blank(
                                qr_payload=cap_qr_payload,
                                tour_number=tour_number,
                                task_number=int(cap_task_num),
                                mode="Задание для капитанов",
                            )
                            zf.writestr(f"{cap_folder}/задание_{cap_task_num}.docx", cap_docx)
                            added_files += 1
                            for extra_i in range(1, 6):
                                cap_extra = word_generator.generate_answer_blank(
                                    qr_payload=cap_qr_payload,
                                    tour_number=tour_number,
                                    task_number=int(cap_task_num),
                                    mode="Задание для капитанов",
                                    tour_task=f"{tour_number}/cap/{cap_task_num}/{extra_i}",
                                )
                                zf.writestr(f"{cap_folder}/дополнительные бланки/extra_{cap_task_num}_{extra_i}.docx", cap_extra)
                                added_files += 1
                        except Exception as exc:  # noqa: BLE001
                            admit_errors.append(
                                {
                                    "registration_id": str(reg.id),
                                    "participant": participant.full_name,
                                    "error": f"Captains task tour {tour_number}: {exc}",
                                }
                            )

            # Keep existing generated PDFs for backward compatibility with scan flow.
            if not sheets and attempt.pdf_file_path:
                try:
                    pdf_bytes = storage.download_file(
                        bucket=settings.minio_bucket_sheets,
                        object_name=attempt.pdf_file_path,
                    )
                    legacy_root = f"Личный зачет/{folder}" if has_team_tours else folder
                    zf.writestr(f"{legacy_root}/legacy/primary.pdf", pdf_bytes)
                    added_files += 1
                except Exception:  # noqa: BLE001
                    pass
                continue

            for index, sheet in enumerate(sheets, start=1):
                if not sheet.pdf_file_path:
                    continue
                try:
                    pdf_bytes = storage.download_file(
                        bucket=settings.minio_bucket_sheets,
                        object_name=sheet.pdf_file_path,
                    )
                    legacy_root = f"Личный зачет/{folder}" if has_team_tours else folder
                    zf.writestr(f"{legacy_root}/legacy/{index}_{sheet.kind.value}.pdf", pdf_bytes)
                    added_files += 1
                except Exception:  # noqa: BLE001
                    continue

        # Generate team tour sheets: one set per institution (captain's attempt only).
        if has_team_tours:
            team_tour_list = [t for t in tours if str(t["mode"]) == "team"]
            # Build mapping: institution_slug → (captain_attempt_id, institution_label)
            captain_attempts: dict[str, tuple] = {}
            for reg in registrations:
                p = reg.participant
                if not p or not reg.attempts:
                    continue
                if not getattr(p, "is_captain", False):
                    continue
                inst_name = _derive_team_name(p)
                inst_slug = _slugify_folder_name(inst_name)
                if inst_slug not in captain_attempts:
                    captain_attempts[inst_slug] = (reg.attempts[0].id, inst_name)

            for inst_slug, (cap_attempt_id, inst_label) in captain_attempts.items():
                team_folder = f"Командный зачет/{inst_slug}"
                for tour in team_tour_list:
                    tour_number = int(tour["tour_number"])
                    mode_label = mode_labels.get(str(tour["mode"]), str(tour["mode"]))
                    task_numbers = tour["task_numbers"]

                    cover_qr_payload = f"attempt:{cap_attempt_id}:tour:{tour_number}:cover"
                    try:
                        cover_docx = word_generator.generate_a3_cover(
                            qr_payload=cover_qr_payload,
                            tour_number=tour_number,
                            mode=mode_label,
                        )
                        zf.writestr(f"{team_folder}/tour_{tour_number}/A3_tour_{tour_number}.docx", cover_docx)
                        added_files += 1
                    except Exception as exc:  # noqa: BLE001
                        admit_errors.append(
                            {
                                "registration_id": "",
                                "participant": inst_label,
                                "error": f"Team A3 tour {tour_number}: {exc}",
                            }
                        )

                    for task_number in task_numbers:
                        task_qr_payload = f"attempt:{cap_attempt_id}:tour:{tour_number}:task:{task_number}"
                        try:
                            task_docx = word_generator.generate_answer_blank(
                                qr_payload=task_qr_payload,
                                tour_number=tour_number,
                                task_number=int(task_number),
                                mode=mode_label,
                            )
                            task_folder_path = f"{team_folder}/tour_{tour_number}/task_{task_number}"
                            zf.writestr(f"{task_folder_path}/task_{task_number}.docx", task_docx)
                            added_files += 1
                            for extra_i in range(1, 6):
                                extra_docx = word_generator.generate_answer_blank(
                                    qr_payload=task_qr_payload,
                                    tour_number=tour_number,
                                    task_number=int(task_number),
                                    mode=mode_label,
                                    tour_task=f"{tour_number}/{task_number}/{extra_i}",
                                )
                                zf.writestr(
                                    f"{task_folder_path}/дополнительные бланки/extra_{extra_i}.docx",
                                    extra_docx,
                                )
                                added_files += 1
                        except Exception as exc:  # noqa: BLE001
                            admit_errors.append(
                                {
                                    "registration_id": "",
                                    "participant": inst_label,
                                    "error": f"Team task {tour_number}/{task_number}: {exc}",
                                }
                            )

    zip_buffer.seek(0)

    headers = {
        "Content-Disposition": f'attachment; filename="special_olympiad_{competition_id}.zip"',
        "X-Admitted-Now": str(admitted_now),
        "X-Archive-Files": str(added_files),
        "X-Admit-Errors": json.dumps(admit_errors),
    }
    return StreamingResponse(zip_buffer, media_type="application/zip", headers=headers)


# ══════════════════════════════════════════════════════════════════════════════
# Badge Template endpoints (JSON-based visual editor)
# ══════════════════════════════════════════════════════════════════════════════

@router.get("/competitions/{competition_id}/badge-template")
async def get_badge_template(
    competition_id: UUID,
    current_user: Annotated[Any, Depends(require_role(UserRole.ADMIN))],
    db: AsyncSession = Depends(get_db),
):
    """Return badge template config for a competition (no background bytes)."""
    result = await db.execute(
        select(BadgeTemplateModel).where(
            BadgeTemplateModel.competition_id == competition_id
        )
    )
    tmpl = result.scalar_one_or_none()
    if tmpl is None:
        return {
            "id": None,
            "competition_id": str(competition_id),
            "config_json": {
                "width_mm": 90,
                "height_mm": 120,
                "background_w_mm": 90,
                "background_h_mm": 120,
                "elements": [],
            },
            "print_per_page": 4,
            "has_background": False,
            "created_at": None,
            "updated_at": None,
        }
    return {
        "id": str(tmpl.id),
        "competition_id": str(tmpl.competition_id),
        "config_json": tmpl.config_json,
        "print_per_page": tmpl.print_per_page,
        "has_background": tmpl.background_image_bytes is not None,
        "created_at": tmpl.created_at.isoformat() if tmpl.created_at else None,
        "updated_at": tmpl.updated_at.isoformat() if tmpl.updated_at else None,
    }


@router.post("/competitions/{competition_id}/badge-template", status_code=200)
async def upsert_badge_template(
    competition_id: UUID,
    body: dict,
    current_user: Annotated[Any, Depends(require_role(UserRole.ADMIN))],
    db: AsyncSession = Depends(get_db),
):
    """Create or update badge template config for a competition.

    Body: {config_json: {...}, print_per_page: 4|6}
    """
    config_json = body.get("config_json", {})
    print_per_page = int(body.get("print_per_page", 4))
    if print_per_page not in (4, 6):
        print_per_page = 4

    result = await db.execute(
        select(BadgeTemplateModel).where(
            BadgeTemplateModel.competition_id == competition_id
        )
    )
    tmpl = result.scalar_one_or_none()

    if tmpl is None:
        import uuid as _uuid
        tmpl = BadgeTemplateModel(
            id=_uuid.uuid4(),
            competition_id=competition_id,
            config_json=config_json,
            print_per_page=print_per_page,
        )
        db.add(tmpl)
    else:
        tmpl.config_json = config_json
        tmpl.print_per_page = print_per_page
        tmpl.updated_at = datetime.utcnow()

    await db.commit()
    await db.refresh(tmpl)
    return {
        "id": str(tmpl.id),
        "competition_id": str(tmpl.competition_id),
        "config_json": tmpl.config_json,
        "print_per_page": tmpl.print_per_page,
        "has_background": tmpl.background_image_bytes is not None,
        "created_at": tmpl.created_at.isoformat() if tmpl.created_at else None,
        "updated_at": tmpl.updated_at.isoformat() if tmpl.updated_at else None,
    }


@router.post("/competitions/{competition_id}/badge-template/background", status_code=200)
async def upload_badge_template_background(
    competition_id: UUID,
    current_user: Annotated[Any, Depends(require_role(UserRole.ADMIN))],
    db: AsyncSession = Depends(get_db),
    file: UploadFile = File(...),
):
    """Upload a background image for the badge template."""
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Файл должен быть изображением")

    img_bytes = await file.read()

    result = await db.execute(
        select(BadgeTemplateModel).where(
            BadgeTemplateModel.competition_id == competition_id
        )
    )
    tmpl = result.scalar_one_or_none()

    if tmpl is None:
        import uuid as _uuid
        tmpl = BadgeTemplateModel(
            id=_uuid.uuid4(),
            competition_id=competition_id,
            config_json={
                "width_mm": 90,
                "height_mm": 120,
                "background_w_mm": 90,
                "background_h_mm": 120,
                "elements": [],
            },
            print_per_page=4,
            background_image_bytes=img_bytes,
        )
        db.add(tmpl)
    else:
        tmpl.background_image_bytes = img_bytes
        tmpl.updated_at = datetime.utcnow()

    await db.commit()
    return {"status": "ok", "size": len(img_bytes)}


@router.get("/competitions/{competition_id}/badge-template/background")
async def get_badge_template_background(
    competition_id: UUID,
    current_user: Annotated[Any, Depends(require_role(UserRole.ADMIN))],
    db: AsyncSession = Depends(get_db),
):
    """Stream the background image bytes."""
    result = await db.execute(
        select(BadgeTemplateModel).where(
            BadgeTemplateModel.competition_id == competition_id
        )
    )
    tmpl = result.scalar_one_or_none()
    if tmpl is None or not tmpl.background_image_bytes:
        raise HTTPException(status_code=404, detail="Фоновое изображение не найдено")

    return StreamingResponse(
        io.BytesIO(tmpl.background_image_bytes),
        media_type="image/png",
        headers={"Cache-Control": "no-cache"},
    )


@router.delete("/competitions/{competition_id}/badge-template", status_code=200)
async def delete_badge_template(
    competition_id: UUID,
    current_user: Annotated[Any, Depends(require_role(UserRole.ADMIN))],
    db: AsyncSession = Depends(get_db),
):
    """Delete the badge template for a competition."""
    result = await db.execute(
        select(BadgeTemplateModel).where(
            BadgeTemplateModel.competition_id == competition_id
        )
    )
    tmpl = result.scalar_one_or_none()
    if tmpl is None:
        raise HTTPException(status_code=404, detail="Шаблон не найден")
    await db.delete(tmpl)
    await db.commit()
    return {"status": "deleted"}


# ── Font serving for badge editor preview ─────────────────────────────────────

_FONTS_DIR = Path(__file__).resolve().parents[5] / "templates" / "word" / "fonts"


@router.get("/badge-fonts")
async def list_badge_fonts(
    current_user: Annotated[Any, Depends(require_role(UserRole.ADMIN))],
):
    """List available custom font filenames for badge editor."""
    if not _FONTS_DIR.exists():
        return []
    return [
        f.name
        for f in sorted(_FONTS_DIR.iterdir())
        if f.suffix.lower() in (".ttf", ".otf")
    ]


@router.get("/badge-fonts/{filename}")
async def get_badge_font(
    filename: str,
    current_user: Annotated[Any, Depends(require_role(UserRole.ADMIN))],
):
    """Serve a custom font file for the badge editor preview."""
    import re as _re
    if not _re.match(r'^[\w\- ]+\.(ttf|otf)$', filename, _re.IGNORECASE):
        raise HTTPException(status_code=400, detail="Недопустимое имя файла")
    font_path = _FONTS_DIR / filename
    if not font_path.exists() or font_path.parent.resolve() != _FONTS_DIR.resolve():
        raise HTTPException(status_code=404, detail="Шрифт не найден")
    suffix = font_path.suffix.lower()
    media_type = "font/otf" if suffix == ".otf" else "font/ttf"
    return StreamingResponse(
        io.BytesIO(font_path.read_bytes()),
        media_type=media_type,
        headers={"Cache-Control": "public, max-age=86400"},
    )


# --- Competition Staff Access Management ---

@router.get("/competitions/{competition_id}/staff", response_model=CompetitionStaffList)
async def list_competition_staff(
    competition_id: UUID,
    current_user: Annotated[User, Depends(require_role(UserRole.ADMIN))],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """List all staff users assigned to a competition."""
    access_repo = UserCompetitionAccessRepositoryImpl(db)
    user_repo = UserRepositoryImpl(db)

    comp_repo = CompetitionRepositoryImpl(db)
    competition = await comp_repo.get_by_id(competition_id)
    if not competition:
        raise HTTPException(status_code=404, detail="Олимпиада не найдена")

    assignments = await access_repo.get_users_for_competition(competition_id)
    items: list[CompetitionStaffItem] = []
    for assignment in assignments:
        user = await user_repo.get_by_id(assignment.user_id)
        if user:
            items.append(
                CompetitionStaffItem(
                    user_id=user.id,
                    email=user.email,
                    role=user.role,
                    assigned_at=assignment.assigned_at,
                )
            )
    return CompetitionStaffList(items=items, total=len(items))


@router.post("/competitions/{competition_id}/staff", response_model=CompetitionStaffItem, status_code=status.HTTP_201_CREATED)
async def assign_competition_staff(
    competition_id: UUID,
    body: AssignStaffRequest,
    current_user: Annotated[User, Depends(require_role(UserRole.ADMIN))],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Assign a staff user (ADMITTER/SCANNER/INVIGILATOR) to a competition."""
    comp_repo = CompetitionRepositoryImpl(db)
    competition = await comp_repo.get_by_id(competition_id)
    if not competition:
        raise HTTPException(status_code=404, detail="Олимпиада не найдена")

    user_repo = UserRepositoryImpl(db)
    user = await user_repo.get_by_id(body.user_id)
    if not user:
        raise HTTPException(status_code=404, detail="Пользователь не найден")
    if not user.role.is_staff or user.role == UserRole.ADMIN:
        raise HTTPException(
            status_code=400,
            detail="Доступ можно выдать только пользователям с ролью ADMITTER, SCANNER или INVIGILATOR",
        )

    access_repo = UserCompetitionAccessRepositoryImpl(db)
    assignment = await access_repo.assign(
        user_id=user.id,
        competition_id=competition_id,
        assigned_by=current_user.id,
    )
    return CompetitionStaffItem(
        user_id=user.id,
        email=user.email,
        role=user.role,
        assigned_at=assignment.assigned_at,
    )


@router.delete("/competitions/{competition_id}/staff/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
async def revoke_competition_staff(
    competition_id: UUID,
    user_id: UUID,
    current_user: Annotated[User, Depends(require_role(UserRole.ADMIN))],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    """Revoke a staff user's access to a competition."""
    access_repo = UserCompetitionAccessRepositoryImpl(db)
    deleted = await access_repo.revoke(user_id=user_id, competition_id=competition_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Назначение не найдено")
